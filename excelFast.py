import sys
import pandas as pd
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import openpyxl
import os
import re
import win32com.client as win32
from tkinter import Tk, messagebox, ttk
import winreg
import concurrent.futures
import datetime
import pythoncom  # 添加pythoncom导入

import matplotlib

matplotlib.use('TkAgg')
import mmap
import io

from functools import lru_cache
import warnings
import gc
import time  # 确保time模块可用

# 抑制不必要的警告
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

# 定义全局缓存
WORKBOOK_CACHE = {}
PARSED_DATA_CACHE = {}


class LoginDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        # 基础窗口设置
        self.title("登录")
        self.geometry("320x450")
        self.configure(bg='#f7f7f7')
        self.resizable(False, False)

        # 样式配置
        self.style = ttk.Style()
        self.style.configure(
            'Custom.TEntry',
            fieldbackground='white',
            borderwidth=0,
            relief='flat'
        )

        self.style.configure(
            'Custom.TButton',
            padding=(0, 8),
            relief='flat',
            background='white',
            borderwidth=0,
            font=('Microsoft YaHei UI', 9)
        )

        # 标题栏
        title_frame = tk.Frame(self, bg='#f7f7f7', height=30)
        title_frame.pack(fill=tk.X)

        # 最小化和关闭按钮
        tk.Button(title_frame, text="−", bg='#f7f7f7', fg='#666666',
                  font=('Microsoft YaHei UI', 11), bd=0,
                  command=self.iconify).pack(side=tk.RIGHT)
        tk.Button(title_frame, text="×", bg='#f7f7f7', fg='#666666',
                  font=('Microsoft YaHei UI', 11), bd=0,
                  command=self.destroy).pack(side=tk.RIGHT, padx=(0, 10))

        # 主标题
        tk.Label(self, text="多文件审计工具",
                 font=('Microsoft YaHei UI', 16),
                 bg='#f7f7f7', fg='#333333').pack(pady=(20, 0))

        # 副标题
        tk.Label(self, text="请登录以继续",
                 font=('Microsoft YaHei UI', 10),
                 bg='#f7f7f7', fg='#666666').pack(pady=(5, 30))

        # 登录表单容器
        form_frame = tk.Frame(self, bg='#f7f7f7')
        form_frame.pack(fill=tk.X, padx=40)

        # 账号输入框
        tk.Label(form_frame, text="账号",
                 font=('Microsoft YaHei UI', 9),
                 bg='#f7f7f7', fg='#666666').pack(anchor='w')
        self.username = tk.Entry(
            form_frame,
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg='#333333',
            insertbackground='#666666',
            relief='flat',
            highlightthickness=1,
            highlightbackground='#e0e0e0',
            highlightcolor='#4a90e2'
        )
        self.username.pack(fill=tk.X, pady=(5, 20))

        # 密码输入框
        tk.Label(form_frame, text="密码",
                 font=('Microsoft YaHei UI', 9),
                 bg='#f7f7f7', fg='#666666').pack(anchor='w')
        self.password = tk.Entry(
            form_frame,
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg='#333333',
            show="●",
            insertbackground='#666666',
            relief='flat',
            highlightthickness=1,
            highlightbackground='#e0e0e0',
            highlightcolor='#4a90e2'
        )
        self.password.pack(fill=tk.X, pady=(5, 30))

        # 登录按钮
        self.login_button = tk.Button(
            form_frame,
            text="登 录",
            font=('Microsoft YaHei UI', 9),
            fg='#666666',
            bg='white',
            activeforeground='#333333',
            activebackground='#f5f5f5',
            relief='flat',
            bd=1,
            highlightthickness=1,
            highlightbackground='#e0e0e0',
            command=self.login
        )
        self.login_button.pack(fill=tk.X, ipady=8)

        # 版权信息
        tk.Label(self, text="溪午不闻钟出品，联系方式:g31050417",
                 font=('Microsoft YaHei UI', 8),
                 bg='#f7f7f7', fg='#999999').pack(side=tk.BOTTOM, pady=20)

        # 窗口拖动
        title_frame.bind('<Button-1>', self.start_move)
        title_frame.bind('<B1-Motion>', self.on_move)

        # 窗口居中
        self.center_window()

        # 登录结果标志
        self.result = False

        # 初始化注册表
        self.registry_path = r"Software\MultiFileAuditTool"
        self.initialize_registry()

    def initialize_registry(self):
        """初始化注册表"""
        try:
            # 尝试创建或打开注册表键
            self.reg_key = winreg.CreateKeyEx(
                winreg.HKEY_CURRENT_USER,
                self.registry_path,
                0,
                winreg.KEY_ALL_ACCESS
            )

            try:
                # 尝试读取登录次数
                self.login_count = winreg.QueryValueEx(self.reg_key, "LoginCount")[0]
            except WindowsError:
                # 如果没有登录次数记录，初始化为0
                self.login_count = 0
                winreg.SetValueEx(
                    self.reg_key,
                    "LoginCount",
                    0,
                    winreg.REG_DWORD,
                    self.login_count
                )
        except WindowsError as e:
            messagebox.showerror("错误", f"无法访问注册表: {str(e)}")
            self.destroy()

    def update_login_count(self):
        """更新登录次数"""
        try:
            self.login_count += 1
            winreg.SetValueEx(
                self.reg_key,
                "LoginCount",
                0,
                winreg.REG_DWORD,
                self.login_count
            )
        except WindowsError as e:
            messagebox.showerror("错误", f"无法更新登录次数: {str(e)}")

    def start_move(self, event):
        self.x = event.x
        self.y = event.y

    def on_move(self, event):
        deltax = event.x - self.x
        deltay = event.y - self.y
        x = self.winfo_x() + deltax
        y = self.winfo_y() + deltay
        self.geometry(f"+{x}+{y}")

    def center_window(self):
        """窗口居中显示"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'+{x}+{y}')

    def login(self):
        """登录验证"""
        username = self.username.get()
        password = self.password.get()

        # 第一套密码（限制使用8次）
        if username == "user" and password == "123456":
            if self.login_count >= 9:
                messagebox.showerror(
                    "错误",
                    "试用结束"
                )
                return
            self.update_login_count()
            remaining = 8 - self.login_count
            messagebox.showinfo(
                "登录成功",
                f"欢迎使用多文件审计工具\n您还可以使用 {remaining} 次"
            )
            self.result = True
            self.destroy()

        # 第二套密码（无限制使用）
        if username == "root" and password == "111":

            self.result = True
            self.destroy()

        else:
            messagebox.showerror("错误", "用户名或密码错误")

    def __del__(self):
        """确保关闭注册表键"""
        try:
            winreg.CloseKey(self.reg_key)
        except:
            pass


class ExcelLoader(threading.Thread):
    def __init__(self, files, progress_callback, finished_callback, error_callback):
        super().__init__()
        self.files = files
        self.progress_callback = progress_callback
        self.finished_callback = finished_callback
        self.error_callback = error_callback
        self.workbook_cache = {}

    def run(self):
        combined_data = {}
        total_files = len(self.files)

        with concurrent.futures.ThreadPoolExecutor(max_workers=min(4, total_files)) as executor:
            future_to_file = {executor.submit(self.process_file, file): file for file in self.files}

            completed = 0
            for future in concurrent.futures.as_completed(future_to_file):
                file = future_to_file[future]
                try:
                    file_data = future.result()
                    if file_data:
                        for sheet_name, data in file_data.items():
                            if sheet_name not in combined_data:
                                combined_data[sheet_name] = []
                            combined_data[sheet_name].append((file, sheet_name))

                    completed += 1
                    self.progress_callback(int(completed / total_files * 100))

                except Exception as e:
                    self.error_callback(f"读取文件 '{file}' 时出错: {str(e)}")

                if file in self.workbook_cache:
                    del self.workbook_cache[file]

        self.finished_callback(combined_data)

    def process_file(self, file):
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=False)
            self.workbook_cache[file] = wb

            result = {}
            for sheet_name in wb.sheetnames:
                result[sheet_name] = True

            return result

        except Exception as e:
            self.error_callback(f"处理文件 '{file}' 时出错: {str(e)}")
            return None
        finally:
            if file in self.workbook_cache:
                self.workbook_cache[file].close()


class EditableTreeview(ttk.Treeview):
    def __init__(self, master, **kw):
        kw['show'] = 'headings'  # 在创建时就设置show属性
        super().__init__(master, **kw)

        self.bind('<Double-1>', self.on_double_click)
        self.bind('<Button-1>', self.on_click)
        self.bind('<Button-3>', self.on_right_click)

        self.entry = None
        self.modified_cells = {}
        self.editing = False  # 添加一个标志来跟踪是否正在编辑
        self.current_item = None  # 当前正在编辑的项
        self.current_column = None  # 当前正在编辑的列

    def set_column_titles(self):
        """设置列标题（@、A、B、C...）"""
        print("设置列标题")
        for i, col in enumerate(self['columns']):
            if i == 0:
                print("设置了@")
                self.heading(col, text='@')  # 第一列为@
            else:
                self.heading(col, text=chr(63 + i))  # A从65开始

    def update_table(self, df):
        # 清除现有数据
        for item in self.get_children():
            self.delete(item)

        # 设置列
        columns = ['文件名'] + [chr(65 + i) for i in range(len(df.columns) - 1)]  # A, B, C...
        self['columns'] = columns

        # 配置每一列
        for col in columns:
            self.column(col, width=100, stretch=True)

        # 设置列标题
        self.set_column_titles()

        # 添加数据行
        for i in range(0, len(df), 2):
            row_id = f'I{i // 2:03d}'
            formula_values = list(df.iloc[i])
            formula_item = self.insert('', 'end', iid=row_id,
                                       values=formula_values,
                                       tags=('formula',))

            if i + 1 < len(df):
                value_values = list(df.iloc[i + 1])
                value_id = f'V{i // 2:03d}'
                self.insert(formula_item, 'end', iid=value_id,
                            values=value_values,
                            tags=('value',))

        # 设置样式
        self.tag_configure('formula', background='#E6F3FF')
        self.tag_configure('value', background='#F0F8FF')

        # 调试信息
        print("Columns:", columns)
        print("Column headings:", [self.heading(col)['text'] for col in columns])

    def on_click(self, event):
        """处理单击事件，展开/折叠值行"""
        item = self.identify('item', event.x, event.y)
        if item:
            # 如果点击的是公式行，切换其子项（值行）的显示状态
            if 'formula' in self.item(item, 'tags'):
                current_state = self.item(item, 'open')
                new_state = not current_state
                self.item(item, open=new_state)  # 切换展开状态
                print(f"切换行 {item} 展开状态为: {new_state}")
                return "break"  # 阻止事件传递，避免选择行干扰展开/折叠

    def on_double_click(self, event):
        """处理双击事件，编辑单元格"""
        region = self.identify("region", event.x, event.y)
        if region == "cell":
            column = self.identify_column(event.x)
            item = self.identify('item', event.x, event.y)
            if item and column and 'formula' in self.item(item, 'tags'):
                # 如果之前有编辑未完成，先保存之前的编辑
                if self.editing and self.entry and self.current_item and self.current_column:
                    self.finish_edit(self.current_item, self.current_column)

                # 开始新的编辑
                self.start_edit(item, column)

    def start_edit(self, item, column):
        x, y, width, height = self.bbox(item, column)

        # 设置当前编辑状态
        self.editing = True
        self.current_item = item
        self.current_column = column

        # 获取当前值
        col_idx = int(column[1:]) - 1
        current_values = self.item(item, 'values')
        if not current_values or col_idx >= len(current_values):
            current_value = ''
        else:
            current_value = current_values[col_idx]

        # 创建编辑框
        if self.entry is None:
            self.entry = ttk.Entry(self)

        self.entry.place(x=x, y=y, width=width, height=height)
        self.entry.delete(0, tk.END)  # 清除上一次的内容
        self.entry.insert(0, current_value)
        self.entry.select_range(0, tk.END)
        self.entry.focus_set()

        # 绑定事件
        self.entry.bind('<Return>', lambda e: self.finish_edit(item, column))
        # 使用after为FocusOut添加延迟，防止和双击冲突
        self.entry.bind('<FocusOut>', lambda e: self.after(10, lambda: self.finish_edit(item, column)))
        # 添加Tab键支持，完成当前编辑并移到下一个单元格
        self.entry.bind('<Tab>', lambda e: self.tab_to_next_cell(item, column))

    def finish_edit(self, item, column):
        # 如果已经不处于编辑状态，直接返回
        if not self.editing or self.entry is None:
            return

        try:
            # 确保控件还存在且可获取值
            new_value = self.entry.get()

            # 标记为非编辑状态
            self.editing = False
            self.current_item = None
            self.current_column = None

            values = list(self.item(item, 'values'))
            col_idx = int(column[1:]) - 1

            if col_idx < 0 or col_idx >= len(values):
                print(f"警告: 无效的列索引 {col_idx}，共有 {len(values)} 列")
                if self.entry:
                    self.entry.destroy()
                    self.entry = None
                return

            old_value = values[col_idx]
            print(
                f"编辑 item={item}, column={column}, col_idx={col_idx}, old_value='{old_value}', new_value='{new_value}'")

            # 无论值是否变化，都保存更改
            # 更新树形表格中的值
            values[col_idx] = new_value
            self.item(item, values=values)

            # 如果是公式行，尝试更新对应的值行显示
            if 'formula' in self.item(item, 'tags'):
                # 获取子项（值行）
                children = self.get_children(item)
                for child in children:
                    if 'value' in self.item(child, 'tags'):
                        # 获取值行的当前值
                        value_values = list(self.item(child, 'values'))
                        if col_idx < len(value_values):
                            # 不再自动使用"计算中..."
                            value_values[col_idx] = new_value
                            self.item(child, values=value_values)
                            print(f"已更新子行 {child} 的值为 '{value_values[col_idx]}'")

            # 记录修改到modified_cells
            if item not in self.modified_cells:
                self.modified_cells[item] = {}
            self.modified_cells[item][col_idx] = new_value
            print(f"已记录修改: 行 {item}, 列 {column}(idx {col_idx}), 旧值: '{old_value}', 新值: '{new_value}'")

            # 通知主窗口有数据被修改
            self.event_generate("<<DataModified>>")
        finally:
            # 确保无论如何都销毁Entry
            if self.entry:
                self.entry.destroy()
                self.entry = None

    def tab_to_next_cell(self, item, column):
        """处理Tab键，完成当前编辑并移到下一个单元格"""
        # 保存当前单元格的编辑
        self.finish_edit(item, column)

        # 确定下一个可编辑单元格
        col_idx = int(column[1:])
        next_col_idx = col_idx + 1

        # 如果有下一列，移到下一列
        if next_col_idx <= len(self['columns']):
            next_column = f'#{next_col_idx}'
            self.after(10, lambda: self.start_edit(item, next_column))
        # 否则移到下一行的第一列
        else:
            # 获取所有可编辑行
            editable_items = [i for i in self.get_children() if 'formula' in self.item(i, 'tags')]
            if item in editable_items:
                current_idx = editable_items.index(item)
                if current_idx + 1 < len(editable_items):
                    next_item = editable_items[current_idx + 1]
                    next_column = '#1'  # 第一列
                    self.after(10, lambda: self.start_edit(next_item, next_column))

        return "break"  # 阻止默认的Tab行为

    def get_modified_data(self):
        return self.modified_cells

    def on_right_click(self, event):
        """处理右键点击事件，显示列操作菜单"""
        column = self.identify_column(event.x)
        if column:
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="批量修改此列",
                             command=lambda: self.batch_edit_column(column))
            menu.add_command(label="批量删除此列内容",
                             command=lambda: self.batch_delete_column(column))
            menu.post(event.x_root, event.y_root)

    def batch_edit_column(self, column):
        """批量修改列数据"""
        # 获取该列第一个公式行的值
        first_formula_item = None
        for item in self.get_children():
            if 'formula' in self.item(item, 'tags'):
                first_formula_item = item
                break

        if first_formula_item:
            current_value = self.item(first_formula_item, 'values')[int(column.replace('#', '')) - 1]
        else:
            current_value = ''

        # 创建对话框
        dialog = tk.Toplevel(self)
        dialog.title("批量修改列数据")
        dialog.geometry("300x230")
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text="请选择修改方式:").pack(pady=5)

        # 单选框：选择修改模式
        mode_var = tk.StringVar(value="replace")
        ttk.Radiobutton(dialog, text="修改为统一值", variable=mode_var, value="replace").pack()
        ttk.Radiobutton(dialog, text="在原公式基础上添加内容", variable=mode_var, value="append").pack()

        tk.Label(dialog, text="请输入值:").pack(pady=5)
        entry = ttk.Entry(dialog)

        # 根据选择的模式来决定输入框的内容
        def update_entry():
            if mode_var.get() == "replace":
                entry.delete(0, tk.END)
                entry.insert(0, current_value)  # 设置当前值
            else:
                entry.delete(0, tk.END)  # 清空输入框

        # 初始时更新输入框内容
        update_entry()

        entry.pack(pady=5)
        entry.select_range(0, tk.END)  # 选中全部文本
        entry.focus_set()  # 设置焦点

        # 监听模式变化时更新输入框内容
        mode_var.trace("w", lambda *args: update_entry())

        def apply_changes():
            new_value = entry.get()
            col_idx = int(column.replace('#', '')) - 1

            # 根据选择的模式，进行不同的操作
            if mode_var.get() == "replace":
                # 如果是替换模式，统一替换所有公式
                for item in self.get_children():
                    if 'formula' in self.item(item, 'tags'):
                        values = list(self.item(item, 'values'))
                        old_value = values[col_idx]

                        if new_value != str(old_value):
                            values[col_idx] = new_value
                            self.item(item, values=values)

                            if item not in self.modified_cells:
                                self.modified_cells[item] = {}
                            self.modified_cells[item][col_idx] = new_value
            elif mode_var.get() == "append":
                # 如果是追加模式，在原公式基础上添加内容
                append_value = new_value  # 获取用户输入的追加内容
                for item in self.get_children():
                    if 'formula' in self.item(item, 'tags'):
                        values = list(self.item(item, 'values'))
                        old_value = values[col_idx]

                        if old_value:  # 只要有值就处理
                            if old_value.startswith("="):
                                # 如果是公式，直接在后面添加
                                new_formula = f"{old_value}{append_value}"
                            else:
                                # 如果是数字，转换成公式形式
                                new_formula = f"={old_value}{append_value}"

                            values[col_idx] = new_formula
                            self.item(item, values=values)

                            if item not in self.modified_cells:
                                self.modified_cells[item] = {}
                            self.modified_cells[item][col_idx] = new_formula
            dialog.destroy()
            messagebox.showinfo("成功", "列数据已批量修改")

        # 添加确认和取消按钮
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="确定", command=apply_changes).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def batch_delete_column(self, column):
        """批量删除列数据"""
        # 获取该列第一个公式行的值作为参考
        first_formula_item = None
        for item in self.get_children():
            if 'formula' in self.item(item, 'tags'):
                first_formula_item = item
                break

        if first_formula_item:
            current_value = self.item(first_formula_item, 'values')[int(column.replace('#', '')) - 1]
        else:
            current_value = ''

        # 创建对话框
        dialog = tk.Toplevel(self)
        dialog.title("批量删除列数据")
        dialog.geometry("300x230")
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text="请选择修改方式:").pack(pady=5)

        # 单选框：选择修改模式
        mode_var = tk.StringVar(value="append")
        # ttk.Radiobutton(dialog, text="修改为统一值", variable=mode_var, value="replace").pack()
        ttk.Radiobutton(dialog, text="在原公式基础上删除内容", variable=mode_var, value="append").pack()

        tk.Label(dialog, text="请输入要删除的内容:").pack(pady=5)
        entry = ttk.Entry(dialog)

        # 根据选择的模式来决定输入框的内容
        def update_entry():
            if mode_var.get() == "replace":
                entry.delete(0, tk.END)

            else:
                entry.delete(0, tk.END)  # 清空输入框

        # 初始时更新输入框内容
        update_entry()

        entry.pack(pady=5)
        entry.select_range(0, tk.END)  # 选中全部文本
        entry.focus_set()  # 设置焦点

        # 监听模式变化时更新输入框内容
        mode_var.trace("w", lambda *args: update_entry())

        def apply_changes():
            value_to_delete = entry.get().strip()
            col_idx = int(column.replace('#', '')) - 1

            # 根据选择的模式，进行不同的操作
            if mode_var.get() == "replace":
                # 如果是替换模式，统一删除所有匹配的内容
                for item in self.get_children():
                    if 'formula' in self.item(item, 'tags'):
                        values = list(self.item(item, 'values'))
                        current_value = str(values[col_idx]) if values[col_idx] is not None else ""

                        if current_value == value_to_delete:
                            values[col_idx] = ""  # 删除内容（设置为空字符串）
                            self.item(item, values=values)

                            if item not in self.modified_cells:
                                self.modified_cells[item] = {}
                            self.modified_cells[item][col_idx] = ""

            elif mode_var.get() == "append":
                # 如果是追加模式，删除追加的内容
                delete_value = value_to_delete  # 获取用户输入的要删除的内容
                for item in self.get_children():
                    if 'formula' in self.item(item, 'tags'):
                        values = list(self.item(item, 'values'))
                        current_value = str(values[col_idx]) if values[col_idx] is not None else ""

                        if current_value:  # 只要有值就处理
                            if current_value.endswith(delete_value):
                                # 如果当前值以要删除的内容结尾，则删除它
                                new_value = current_value[:-len(delete_value)]
                                values[col_idx] = new_value
                                self.item(item, values=values)

                                if item not in self.modified_cells:
                                    self.modified_cells[item] = {}
                                self.modified_cells[item][col_idx] = new_value

            dialog.destroy()
            messagebox.showinfo("成功", "列数据已批量修改")

        # 添加确认和取消按钮
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="确定", command=apply_changes).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)


class AuditTool(tk.Tk):
    def __init__(self):
        # 在初始化时先检查时间
        if not self.check_validity():
            # 显示一个普通的错误信息，不暴露真实原因
            messagebox.showerror("错误", "程序初始化失败")
            sys.exit(1)

        super().__init__()

        self.fast_mode = tk.BooleanVar(value=True)  # 默认启用快速模式
        self.exclude_hidden = tk.BooleanVar(value=True)  # 默认排

        self.title("多文件审计工具")
        self.geometry("1200x800")
        self.configure(bg='#f7f7f7')  # 设置窗口背景色

        # 配置样式
        self.style = ttk.Style()
        self.style.theme_use('clam')

        # 按钮样式
        self.style.configure(
            'Custom.TButton',
            padding=(12, 6),
            relief='flat',
            background='#ffffff',
            foreground='#666666',
            borderwidth=0,
            focuscolor='#ffffff',
            font=('Microsoft YaHei UI', 9)
        )

        # Frame样式
        self.style.configure(
            'Custom.TFrame',
            background='#f7f7f7',
            borderwidth=0
        )

        # LabelFrame样式
        self.style.configure(
            'Custom.TLabelframe',
            background='#ffffff',
            borderwidth=1,
            relief='solid',
            bordercolor='#eeeeee'
        )

        self.style.configure(
            'Custom.TLabelframe.Label',
            background='#ffffff',
            foreground='#666666',
            font=('Microsoft YaHei UI', 9)
        )

        # Entry和Combobox样式
        self.style.configure(
            'TEntry',
            fieldbackground='#ffffff',
            borderwidth=1,
            relief='solid',
            bordercolor='#eeeeee'
        )

        self.style.configure(
            'TCombobox',
            fieldbackground='#ffffff',
            background='#ffffff',
            borderwidth=1,
            arrowsize=12
        )

        # Treeview样式
        self.style.configure(
            'Custom.Treeview',
            background='#ffffff',
            fieldbackground='#ffffff',
            foreground='#333333',
            borderwidth=1,
            relief='flat',
            font=('Microsoft YaHei UI', 9)
        )

        self.style.configure(
            'Custom.Treeview.Heading',
            background='#f9f9f9',
            foreground='#666666',
            relief='flat',
            borderwidth=1,
            font=('Microsoft YaHei UI', 9)
        )

        # 移除表格的焦点边框
        self.style.layout('Custom.Treeview', [
            ('Custom.Treeview.treearea', {'sticky': 'nswe'})
        ])

        # 进度条样式
        self.style.configure(
            'Custom.Horizontal.TProgressbar',
            troughcolor='#f0f0f0',
            background='#4a90e2',
            bordercolor='#f0f0f0',
            lightcolor='#4a90e2',
            darkcolor='#4a90e2',
            borderwidth=0,
            thickness=6
        )

        # 登录逻辑保持不变
        self.withdraw()
        self.login_dialog = LoginDialog(self)
        self.wait_window(self.login_dialog)

        if self.login_dialog.result:
            self.deiconify()
            self.create_widgets()
        else:
            self.destroy()

    def check_validity(self):
        """检查程序有效期"""
        try:
            # 获取当前时间
            current_time = datetime.datetime.now()

            # 设置失效时间（2024年11月31日）
            expiry_time = datetime.datetime(2042, 12, 30, 23, 59, 59)

            # 检查是否系统时间被修改
            if self.is_time_tampered():
                return False

            # 检查是否超过有效期
            return current_time <= expiry_time

        except:
            return False

    def is_time_tampered(self):
        """检查系统时间是否被篡改"""
        try:
            # 尝试从网络获取时间
            import ntplib
            client = ntplib.NTPClient()
            response = client.request('pool.ntp.org')
            network_time = datetime.datetime.fromtimestamp(response.tx_time)

            # 获取本地时间
            local_time = datetime.datetime.now()

            # 如果本地时间和网络时间差异超过1小时，认为时间被篡改
            time_diff = abs((network_time - local_time).total_seconds())
            return time_diff > 3600

        except:
            # 如果无法获取网络时间，则仅使用本地时间
            return False

    def reload_data(self):
        if self.filtered_data is not None:
            self.load_all_data()
            self.filter_data()
        print("数据已重新加载")

    def create_widgets(self):
        # 主容器
        main_frame = ttk.Frame(self, style='Custom.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # 顶部按钮区域
        top_frame = ttk.Frame(main_frame, style='Custom.TFrame')
        top_frame.pack(fill=tk.X, pady=(0, 10))

        # 按钮
        self.load_button = ttk.Button(
            top_frame,
            text="加载 Excel 文件",
            style='Custom.TButton',
            command=self.load_excel_files
        )
        self.load_button.pack(side=tk.LEFT, padx=(0, 8))

        self.clear_button = ttk.Button(
            top_frame,
            text="清除所有文件",
            style='Custom.TButton',
            command=self.clear_files
        )
        self.clear_button.pack(side=tk.LEFT, padx=(0, 8))

        # 添加规则审计按钮
        self.audit_button = ttk.Button(
            top_frame,
            text="规则审计",
            style='Custom.TButton',
            command=self.audit_rules
        )
        self.audit_button.pack(side=tk.LEFT, padx=(0, 8))

        # 添加预计算按钮
        self.precalculate_button = ttk.Button(
            top_frame,
            text="搜索项为计算值时使用-单独执行版",
            style='Custom.TButton',
            command=self.precalculate_excel_files
        )
        self.precalculate_button.pack(side=tk.LEFT)

        # 添加进度条
        self.progress_bar = ttk.Progressbar(
            top_frame,
            length=200,
            mode='determinate',
            style='Custom.Horizontal.TProgressbar'
        )
        self.progress_bar.pack(side=tk.RIGHT, padx=(0, 5))

        # 内容区域
        content_frame = ttk.Frame(main_frame, style='Custom.TFrame')
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 左侧文件列表
        left_frame = ttk.LabelFrame(
            content_frame,
            text="已加载的文件",
            style='Custom.TLabelframe'
        )
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        self.file_list = tk.Listbox(
            left_frame,
            background='white',
            selectmode=tk.EXTENDED,
            relief='flat',
            borderwidth=0,
            highlightthickness=1,
            highlightbackground='#eeeeee'
        )
        self.file_list.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 右侧操作区域
        right_frame = ttk.LabelFrame(
            content_frame,
            text="操作区域",
            style='Custom.TLabelframe'
        )
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 操作区域内的控件
        options_frame = ttk.Frame(right_frame, style='Custom.TFrame')
        options_frame.pack(fill=tk.X, padx=10, pady=5)

        # 表格选择
        sheet_frame = ttk.Frame(options_frame, style='Custom.TFrame')
        sheet_frame.pack(fill=tk.X, pady=2)
        ttk.Label(sheet_frame, text="选择表:", style='Custom.TLabel').pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(sheet_frame)
        self.sheet_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))

        # 关键词输入
        keyword_frame = ttk.Frame(options_frame, style='Custom.TFrame')
        keyword_frame.pack(fill=tk.X, pady=2)
        ttk.Label(keyword_frame, text="输入关键词:", style='Custom.TLabel').pack(side=tk.LEFT)
        self.keyword_input = ttk.Entry(keyword_frame)
        self.keyword_input.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))

        # 添加匹配模式选择
        match_frame = ttk.Frame(options_frame, style='Custom.TFrame')
        match_frame.pack(fill=tk.X, pady=2)
        ttk.Label(match_frame, text="匹配模式:", style='Custom.TLabel').pack(side=tk.LEFT)
        self.match_mode = tk.StringVar(value="exact")  # 默认精准匹配
        ttk.Radiobutton(match_frame, text="精准匹配", variable=self.match_mode,
                        value="exact").pack(side=tk.LEFT, padx=(5, 10))
        ttk.Radiobutton(match_frame, text="模糊匹配", variable=self.match_mode,
                        value="fuzzy").pack(side=tk.LEFT)

        # 添加预计算选项
        precalculate_frame = ttk.Frame(options_frame, style='Custom.TFrame')
        precalculate_frame.pack(fill=tk.X, pady=2)
        self.precalculate_before_search = tk.BooleanVar(value=False)  # 默认不在搜索前预计算
        ttk.Checkbutton(precalculate_frame, text="搜索项为计算值时勾选-嵌入搜索流程版",
                        variable=self.precalculate_before_search).pack(anchor="w")

        # 操作按钮
        button_frame = ttk.Frame(options_frame, style='Custom.TFrame')
        button_frame.pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="筛选数据",
                   style='Custom.TButton',
                   command=self.filter_data).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="导出筛选结果",
                   style='Custom.TButton',
                   command=self.export_filtered_data).pack(side=tk.LEFT)

        # 表格区域
        table_frame = ttk.Frame(right_frame, style='Custom.TFrame')
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 添加水平和垂直滚动条
        h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)

        # 创建表格并配置滚动条
        self.table = EditableTreeview(
            table_frame,
            style='Custom.Treeview',
            xscrollcommand=h_scrollbar.set,
            yscrollcommand=v_scrollbar.set
        )

        # 设置滚动条的滚动命令
        h_scrollbar.config(command=self.table.xview)
        v_scrollbar.config(command=self.table.yview)

        # 使用grid布局管理器
        self.table.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')

        # 配置grid权重，使表格能够自适应大小
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # 保存按钮
        save_frame = ttk.Frame(right_frame, style='Custom.TFrame')
        save_frame.pack(fill=tk.X, padx=10, pady=5)
        self.save_button = ttk.Button(
            save_frame,
            text="保存修改",
            style='Custom.TButton',
            command=self.save_changes
        )
        self.save_button.pack(side=tk.RIGHT)

        # 其他初始化代码保持不变
        self.excel_files = []
        self.combined_data = {}
        self.filtered_data = None
        self.cell_locations = []

        # 初始化表格列标题
        self.table.set_column_titles()

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

    def load_excel_files(self):
        print(f"【函数调用】load_excel_files() - 正在加载Excel文件")
        try:
            file_paths = filedialog.askopenfilenames(title="选择 Excel 文件",
                                                     filetypes=[("Excel Files", "*.xlsx *.xls")])
            if file_paths:
                # 检查文件是否已经加载
                new_files = [f for f in file_paths if f not in self.excel_files]
                if new_files:
                    # # 启动一个后台线程来自动处理Excel安全警告
                    # warning_handler_thread = threading.Thread(target=self.handle_excel_warnings, daemon=True)
                    # warning_handler_thread.start()

                    self.excel_files.extend(new_files)
                    self.update_file_list()
                    self.load_all_data()
                else:
                    messagebox.showinfo("提示", "所选文件已经加载")
        except Exception as e:
            print(f"加载文件错误: {str(e)}")
            print(traceback.format_exc())

    def update_file_list(self):
        try:
            self.file_list.delete(0, tk.END)
            for file in self.excel_files:
                self.file_list.insert(tk.END, file)
        except Exception as e:
            print(f"更新文件表错误: {str(e)}")
            print(traceback.format_exc())

    def clear_files(self):
        try:
            self.excel_files.clear()
            self.update_file_list()
            self.combined_data.clear()
            self.sheet_combo['values'] = []
            self.table.delete('1.0', tk.END)
            self.filtered_data = None
            self.cell_locations = []
        except Exception as e:
            print(f"文件错误: {str(e)}")
            print(traceback.format_exc())

    def load_all_data(self):
        print(f"【函数调用】load_all_data() - 正在加载所有数据")
        self.progress_bar['value'] = 0
        self.loader = ExcelLoader(
            self.excel_files,
            self.update_progress,
            self.on_load_finished,
            self.on_load_error
        )
        self.loader.start()

    def update_progress(self, value):
        self.progress_bar['value'] = value

    def on_load_finished(self, data):
        self.combined_data = data
        self.update_sheet_combo()
        messagebox.showinfo("成功", "所有文件加载完成")

    def on_load_error(self, error_msg):
        messagebox.showwarning("警告", error_msg)

    def update_sheet_combo(self):
        try:
            self.sheet_combo['values'] = list(self.combined_data.keys())
        except Exception as e:
            print(f"更新子表下拉框错误: {str(e)}")
            print(traceback.format_exc())

    def refresh_excel_files_parallel(self, file_paths, progress_callback=None):
        """并行处理多个Excel实例，大幅提高速度"""
        # 确定最佳并行数
        max_parallel = min(os.cpu_count() or 4, 4)  # 最多4个Excel实例

        # 分批处理文件
        batches = [file_paths[i:i + len(file_paths) // max_parallel + 1]
                   for i in range(0, len(file_paths), len(file_paths) // max_parallel + 1)]

        # 创建线程池
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_parallel) as executor:
            # 提交所有批次任务
            futures = []
            for batch in batches:
                futures.append(executor.submit(self._process_excel_batch, batch, progress_callback))

            # 等待所有任务完成
            concurrent.futures.wait(futures)

    def _process_excel_batch(self, batch, progress_callback):
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = 0
        excel.DisplayAlerts = 0
        excel.EnableEvents = False

        try:
            for file_path in batch:
                try:
                    wb = excel.Workbooks.Open(file_path, UpdateLinks=False, ReadOnly=False)
                    excel.Calculation = -4105  # xlCalculationAutomatic
                    wb.Close(SaveChanges=True)
                except Exception as e:
                    # 可选：打印错误信息以便调试
                    print(f"处理文件出错: {file_path}, 错误: {e}")
        finally:
            excel.Quit()

    def filter_data(self):
        """极端优化的数据筛选方法 - 采用激进优化策略"""
        try:
            if not self.combined_data:
                messagebox.showwarning("警告", "请先加载Excel文件")
                return

            selected_sheet = self.sheet_combo.get()
            keyword = self.keyword_input.get().strip()

            if not keyword:
                messagebox.showwarning("警告", "请输入关键词")
                return

            if not selected_sheet:
                messagebox.showwarning("警告", "请选择一个子表")
                return

            if selected_sheet not in self.combined_data:
                messagebox.showwarning("警告", f"子表 '{selected_sheet}' 不存在")
                return

            # 创建进度窗口
            progress_window = tk.Toplevel(self)
            progress_window.title("处理中")
            progress_window.geometry("300x200")
            progress_window.transient(self)
            progress_window.grab_set()

            # 更详细的进度显示
            refresh_label = ttk.Label(progress_window, text="正在处理数据...")
            refresh_label.pack(pady=5)
            refresh_progress = ttk.Progressbar(progress_window, length=200, mode='determinate')
            refresh_progress.pack(pady=5)

            search_label = ttk.Label(progress_window, text="正在搜索...")
            search_label.pack(pady=5)
            search_progress = ttk.Progressbar(progress_window, length=200, mode='determinate')
            search_progress.pack(pady=5)

            # 添加详细状态显示
            status_label = ttk.Label(progress_window, text="准备中...")
            status_label.pack(pady=5)

            # 添加状态跟踪
            stats_frame = ttk.Frame(progress_window)
            stats_frame.pack(pady=10, fill='x')

            ttk.Label(stats_frame, text="已处理文件:").grid(row=0, column=0, sticky='w', padx=5)
            files_processed_label = ttk.Label(stats_frame, text="0")
            files_processed_label.grid(row=0, column=1, sticky='w', padx=5)

            ttk.Label(stats_frame, text="匹配记录:").grid(row=1, column=0, sticky='w', padx=5)
            rows_found_label = ttk.Label(stats_frame, text="0")
            rows_found_label.grid(row=1, column=1, sticky='w', padx=5)

            # 预处理关键词
            keyword_lower = keyword.lower()
            match_mode = self.match_mode.get()

            # 确保pandas可用（在这里引入而不是在local作用域内）
            import pandas as pd_local

            # 尝试加载优化库
            try:
                import win32com.client as win32
                import numpy as np
                use_win32 = True
            except:
                use_win32 = False

            # 清理内存，确保有足够空间处理
            gc.collect()

            # 主处理线程
            def process_files():
                # 导入pandas到本地作用域
                import pandas as pd

                try:
                    # 获取需要处理的文件
                    files_to_process = list(set(file for file, _ in self.combined_data[selected_sheet]))
                    total_files = len(files_to_process)

                    # 如果用户选择了搜索前预计算，先执行预计算
                    if hasattr(self, 'precalculate_before_search') and self.precalculate_before_search.get():
                        status_label.config(text="正在预计算公式...")
                        refresh_label.config(text="预计算进度")
                        refresh_progress['value'] = 0
                        progress_window.update_idletasks()

                        # 确定最佳并行数
                        max_workers = min(os.cpu_count() or 4, 4)

                        # 优化批处理分配
                        if len(files_to_process) <= max_workers:
                            batches = [[file] for file in files_to_process]
                        else:
                            batch_size = len(files_to_process) // max_workers
                            remainder = len(files_to_process) % max_workers

                            batches = []
                            start_idx = 0
                            for i in range(max_workers):
                                end_idx = start_idx + batch_size + (1 if i < remainder else 0)
                                batches.append(files_to_process[start_idx:end_idx])
                                start_idx = end_idx

                        processed_count = 0

                        # 使用线程池并行处理预计算
                        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                            futures = []
                            for batch in batches:
                                if batch:
                                    futures.append(executor.submit(self.process_excel_batch, batch))

                            # 等待所有任务完成
                            for future in concurrent.futures.as_completed(futures):
                                try:
                                    batch_processed, batch_errors = future.result()
                                    processed_count += len(batch_processed)

                                    # 更新预计算进度
                                    progress = int(processed_count * 100 / total_files)
                                    refresh_progress['value'] = progress
                                    refresh_label.config(text=f"已预计算 {processed_count}/{total_files} 个文件")
                                    files_processed_label.config(text=str(processed_count))
                                    status_label.config(text=f"预计算中... {progress}%")
                                    progress_window.update_idletasks()

                                except Exception as e:
                                    print(f"预计算批处理出错: {str(e)}")

                        status_label.config(text="预计算完成，开始搜索...")
                        progress_window.update_idletasks()

                    # 计算最佳线程/进程数
                    cpu_count = os.cpu_count() or 4
                    memory_gb = self._get_available_memory_gb()
                    # 根据可用内存和CPU动态调整
                    if memory_gb > 16:
                        optimal_workers = min(cpu_count * 2, 16)  # 高内存机器可以用更多线程
                    elif memory_gb > 8:
                        optimal_workers = min(cpu_count, 8)  # 中等内存
                    else:
                        optimal_workers = max(2, cpu_count // 2)  # 低内存

                    # 更新UI的函数 - 增强版
                    def update_ui(progress=None, text=None, is_refresh=True,
                                  files_processed=None, rows_found=None, status=None):
                        if progress is not None:
                            if is_refresh:
                                self.after(0, lambda: refresh_progress.configure(value=progress))
                            else:
                                self.after(0, lambda: search_progress.configure(value=progress))

                        if text is not None:
                            if is_refresh:
                                self.after(0, lambda: refresh_label.configure(text=text))
                            else:
                                self.after(0, lambda: search_label.configure(text=text))

                        if status is not None:
                            self.after(0, lambda: status_label.configure(text=status))

                        if files_processed is not None:
                            self.after(0, lambda: files_processed_label.configure(text=str(files_processed)))

                        if rows_found is not None:
                            self.after(0, lambda: rows_found_label.configure(text=str(rows_found)))

                    # 处理结果存储
                    filtered_rows = []
                    self.cell_locations = []
                    all_columns = None
                    rows_found_count = 0

                    # 使用缓存加速
                    global WORKBOOK_CACHE
                    global PARSED_DATA_CACHE

                    # 确定处理引擎策略
                    use_memory_mapping = True  # 默认启用内存映射

                    # 获取文件总大小，以确定是否需要分批处理
                    total_size_mb = sum(os.path.getsize(f) / 1024 / 1024 for f in files_to_process if os.path.exists(f))
                    update_ui(status=f"总文件大小: {total_size_mb:.2f} MB, 使用 {optimal_workers} 个工作线程")

                    # 确定批处理大小 - 根据文件大小动态调整
                    if total_size_mb > 1000:  # 大于1GB
                        batch_size = 2
                        use_memory_mapping = False  # 大文件禁用内存映射
                    elif total_size_mb > 500:
                        batch_size = 3
                    elif total_size_mb > 200:
                        batch_size = 5
                    else:
                        batch_size = 8

                    # 只在非快速模式下刷新Excel公式，并且没有使用预计算功能时
                    if not self.fast_mode.get() and not (
                            hasattr(self, 'precalculate_before_search') and self.precalculate_before_search.get()):
                        update_ui(5, "正在刷新Excel公式...", True, 0, 0, "刷新计算")
                        try:
                            self.refresh_excel_files_parallel(files_to_process,
                                                              lambda p: update_ui(p,
                                                                                  f"已刷新 {int(p * len(files_to_process) / 100)}/{len(files_to_process)} 个文件",
                                                                                  True))
                        except Exception as e:
                            print(f"刷新Excel公式时出错: {str(e)}")
                    else:
                        if hasattr(self, 'precalculate_before_search') and self.precalculate_before_search.get():
                            update_ui(5, "已预计算公式，跳过二次计算", True, 0, 0, "使用已计算的值")
                        else:
                            update_ui(5, "快速模式：跳过公式计算", True, 0, 0, "使用缓存值")

                    # 1. 优先使用Win32 COM API极速处理
                    if use_win32:
                        update_ui(5, "正在加载高性能Excel引擎...", status="准备Win32引擎")
                        try:
                            # 使用多引擎策略处理
                            multi_engine_results = self._process_with_multi_engine(
                                files_to_process,
                                selected_sheet,
                                keyword_lower,
                                match_mode,
                                batch_size,
                                optimal_workers,
                                lambda p, t, fp, rf, st: update_ui(p, t, True, fp, rf, st)
                            )

                            # 合并结果
                            if multi_engine_results:
                                all_columns = multi_engine_results.get('columns')
                                filtered_rows = multi_engine_results.get('rows', [])
                                self.cell_locations = multi_engine_results.get('locations', [])
                                rows_found_count = len(filtered_rows) // 2
                        except Exception as e:
                            print(f"多引擎处理错误: {str(e)}")
                            print(traceback.format_exc())
                            update_ui(status=f"多引擎处理失败: {str(e)[:50]}...")

                    # 如果多引擎处理失败，回退到并行线程处理
                    if not filtered_rows:
                        update_ui(10, "使用并行处理引擎...", status="初始化并行处理")

                        # 使用线程池 - 强化版
                        with concurrent.futures.ThreadPoolExecutor(max_workers=optimal_workers) as executor:
                            # 创建分批任务 - 每批最多batch_size个文件
                            tasks = []
                            for i in range(0, len(files_to_process), batch_size):
                                batch_files = files_to_process[i:i + batch_size]
                                task = {
                                    'files': batch_files,
                                    'sheet': selected_sheet,
                                    'keyword': keyword_lower,
                                    'match_mode': match_mode,
                                    'use_memory_mapping': use_memory_mapping
                                }
                                tasks.append(task)

                            # 提交批处理任务
                            futures = {executor.submit(self._process_file_batch, task): i
                                       for i, task in enumerate(tasks)}

                            # 处理完成的任务
                            processed_files = 0
                            for future in concurrent.futures.as_completed(futures):
                                task_idx = futures[future]
                                task = tasks[task_idx]
                                batch_files = task['files']

                                # 更新处理进度
                                processed_files += len(batch_files)
                                progress = int((processed_files / total_files) * 100)

                                # 获取结果
                                try:
                                    batch_result = future.result()

                                    # 提取列信息
                                    if all_columns is None and batch_result and 'columns' in batch_result:
                                        all_columns = batch_result['columns']

                                    # 添加匹配行
                                    if batch_result and 'rows' in batch_result:
                                        for row_data in batch_result['rows']:
                                            filtered_rows.append(row_data['formula_values'])
                                            filtered_rows.append(row_data['value_values'])
                                            self.cell_locations.append(row_data['location'])
                                            rows_found_count += 1

                                    # 更新UI
                                    update_ui(
                                        progress,
                                        f"已处理 {processed_files}/{total_files} 个文件",
                                        False,
                                        processed_files,
                                        rows_found_count,
                                        f"批次 {task_idx + 1}/{len(tasks)} 完成"
                                    )

                                except Exception as batch_e:
                                    print(f"批处理 {task_idx} 出错: {str(batch_e)}")
                                    update_ui(
                                        status=f"处理批次 {task_idx + 1} 出错: {str(batch_e)[:50]}..."
                                    )

                    # 确保有默认列名
                    if all_columns is None:
                        all_columns = ['文件名'] + [f'列 {i}' for i in range(1, 20)]
                    # 打印匹配行数
                    print(f"匹配结果: {rows_found_count} 行")
                    update_ui(status=f"搜索完成，找到 {rows_found_count} 条匹配记录")

                    # 最终结果处理
                    if filtered_rows:
                        # 确保我们捕获异常并正确传递给UI线程
                        try:
                            # 先获取最大列数
                            max_columns = 0
                            for row in filtered_rows:
                                if row:  # 确保行不为None
                                    max_columns = max(max_columns, len(row))

                            # 确保所有行有相同的列数
                            for row in filtered_rows:
                                if row and len(row) < max_columns:
                                    row.extend([None] * (max_columns - len(row)))

                            # 如果all_columns比实际数据列少，则扩展它
                            if len(all_columns) < max_columns:
                                all_columns.extend([f'额外列 {i + 1}' for i in range(len(all_columns), max_columns)])

                            # 或者如果all_columns比实际数据列多，则截断它
                            all_columns = all_columns[:max_columns]

                            # 现在创建DataFrame
                            update_ui(status="创建数据表格中...")
                            self.filtered_data = pd.DataFrame(filtered_rows, columns=all_columns)
                            self.after(0, lambda: self.update_table(self.filtered_data))
                            self.after(100, lambda: self.table.set_column_titles())  # 确保表格列标题正确设置
                            update_ui(status="完成！数据已加载到表格")
                        except Exception as local_e:
                            print(f"创建DataFrame错误: {str(local_e)}")
                            error_msg = str(local_e)
                            self.after(0,
                                       lambda error=error_msg: messagebox.showerror("错误", f"处理数据时出错: {error}"))
                            update_ui(status=f"创建数据表格失败: {str(local_e)[:50]}...")
                    else:
                        self.after(0, lambda: messagebox.showwarning("警告", "没有找到匹配的数据"))
                        update_ui(status="未找到匹配记录")

                except Exception as process_e:
                    print(f"处理过程发生错误: {str(process_e)}")
                    print(traceback.format_exc())
                    error_msg = str(process_e)
                    self.after(0, lambda error=error_msg: messagebox.showerror("错误", f"处理数据时出错: {error}"))
                    update_ui(status=f"处理失败: {str(process_e)[:50]}...")
                finally:
                    # 清理资源
                    gc.collect()
                    self.after(0, lambda: progress_window.destroy())

            # 启动处理线程
            threading.Thread(target=process_files, daemon=True).start()

        except Exception as e:
            print(f"筛选数据错误: {str(e)}")
            print(traceback.format_exc())
            messagebox.showerror("错误", f"筛选数据出错: {str(e)}")

    def _get_available_memory_gb(self) -> float:
        """获取可用内存（GB）"""
        try:
            import psutil
            return psutil.virtual_memory().available / (1024 ** 3)
        except:
            # 如果无法获取，返回一个保守的估计
            return 4.0

    def _process_with_multi_engine(self, files, sheet_name, keyword, match_mode,
                                   batch_size, workers, progress_callback=None):
        """多引擎处理策略 - 同时使用Win32 COM、Pandas和openpyxl"""
        try:
            # 根据文件大小和类型选择最佳引擎
            grouped_files = self._group_files_by_type(files)

            # 初始化结果
            all_columns = None
            all_rows = []
            all_locations = []
            processed_count = 0
            rows_found = 0

            # 使用不同引擎处理不同类型的文件
            engines = ['win32', 'pandas', 'openpyxl']

            for engine in engines:
                if engine not in grouped_files or not grouped_files[engine]:
                    continue

                engine_files = grouped_files[engine]
                if progress_callback:
                    progress_callback(
                        int(processed_count / len(files) * 100),
                        f"使用 {engine} 引擎处理 {len(engine_files)} 个文件",
                        processed_count,
                        rows_found,
                        f"切换到 {engine.upper()} 引擎"
                    )

                # 选择合适的处理方法
                if engine == 'win32':
                    result = self._process_excel_batch_win32(
                        engine_files, sheet_name, keyword, match_mode, batch_size,
                        lambda p, t, f, r, s: progress_callback(
                            int((processed_count + p * len(engine_files) / 100) / len(files) * 100),
                            t, processed_count + f, rows_found + r, s
                        )
                    )
                elif engine == 'pandas':
                    result = self._process_with_pandas(
                        engine_files, sheet_name, keyword, match_mode, batch_size, workers,
                        lambda p, t, f, r, s: progress_callback(
                            int((processed_count + p * len(engine_files) / 100) / len(files) * 100),
                            t, processed_count + f, rows_found + r, s
                        )
                    )
                else:  # openpyxl
                    result = self._process_with_openpyxl(
                        engine_files, sheet_name, keyword, match_mode, batch_size, workers,
                        lambda p, t, f, r, s: progress_callback(
                            int((processed_count + p * len(engine_files) / 100) / len(files) * 100),
                            t, processed_count + f, rows_found + r, s
                        )
                    )

                # 合并结果
                if result:
                    if all_columns is None and 'columns' in result:
                        all_columns = result['columns']

                    if 'rows' in result:
                        all_rows.extend(result['rows'])
                        rows_found += len(result['rows'])

                    if 'locations' in result:
                        all_locations.extend(result['locations'])

                # 更新处理计数
                processed_count += len(engine_files)

            # 返回合并的结果
            return {
                'columns': all_columns,
                'rows': all_rows,
                'locations': all_locations
            }

        except Exception as e:
            print(f"多引擎处理出错: {str(e)}")
            print(traceback.format_exc())
            return None

    def _group_files_by_type(self, files):
        """根据文件特性将文件分组到不同引擎"""
        grouped = {
            'win32': [],  # 大型复杂Excel文件、带有VBA的文件
            'pandas': [],  # 中型常规Excel文件
            'openpyxl': []  # 小型简单Excel文件
        }

        for file in files:
            try:
                # 获取文件大小（MB）
                size_mb = os.path.getsize(file) / (1024 * 1024)

                # 检测文件类型和特性
                if file.endswith('.xls') or size_mb > 10:
                    # 旧格式或大文件使用Win32
                    grouped['win32'].append(file)
                elif size_mb > 2:
                    # 中等大小文件使用pandas
                    grouped['pandas'].append(file)
                else:
                    # 小文件使用openpyxl
                    grouped['openpyxl'].append(file)
            except:
                # 如果无法确定，默认使用openpyxl
                grouped['openpyxl'].append(file)

        return grouped

    def _process_file_batch(self, task):
        """增强的文件批处理函数"""
        # 获取参数
        files = task['files']
        sheet = task['sheet']
        keyword = task['keyword']
        match_mode = task['match_mode']
        use_memory_mapping = task.get('use_memory_mapping', False)

        # # 获取隐藏行列筛选状态
        # exclude_hidden = getattr(self, 'exclude_hidden', None)
        # exclude_hidden = exclude_hidden.get() if exclude_hidden else True
        #
        # print(f"【重要】文件批处理中隐藏内容筛选状态: {exclude_hidden}")

        # 初始化结果
        columns = None
        result_rows = []

        # 处理批次中的每个文件
        for file in files:
            # 检查内存映射模式是否可用
            if use_memory_mapping and file.endswith('.xlsx'):
                try:
                    # 使用内存映射加速
                    with open(file, 'rb') as f:
                        # 创建内存映射文件
                        with mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ) as mmapped_file:
                            # 使用BytesIO作为内存优化的包装器
                            buffer = io.BytesIO(mmapped_file)

                            # 使用openpyxl的read_only加载工作簿
                            wb = openpyxl.load_workbook(buffer, read_only=True, data_only=True)

                            try:
                                # 处理工作簿
                                if sheet in wb.sheetnames:
                                    ws = wb[sheet]

                                    # 查找匹配行
                                    file_result = self._search_in_worksheet_optimized(
                                        ws, file, sheet, keyword, match_mode
                                    )

                                    # 处理结果
                                    if file_result:
                                        if columns is None and 'columns' in file_result:
                                            columns = file_result['columns']

                                        if 'rows' in file_result:
                                            result_rows.extend(file_result['rows'])
                            finally:
                                wb.close()
                        continue  # 成功处理，继续下一个文件
                except:
                    # 内存映射模式失败，回退到标准模式
                    pass

            # 如果内存映射失败或不可用，使用标准方法
            file_result = self._process_single_file(file, sheet, keyword, match_mode)

            # 处理结果
            if file_result:
                if columns is None and 'columns' in file_result:
                    columns = file_result['columns']

                if 'rows' in file_result:
                    result_rows.extend(file_result['rows'])

        # 返回批次结果
        return {
            'columns': columns,
            'rows': result_rows
        }

    @lru_cache(maxsize=32)
    def _search_in_worksheet_optimized(self, worksheet, file, sheet_name, keyword, match_mode):
        """优化的工作表搜索 - 使用缓存和内存优化技术"""

        # # 获取隐藏行列状态
        # exclude_hidden = getattr(self, 'exclude_hidden', None)
        # exclude_hidden = exclude_hidden.get() if exclude_hidden else True
        #
        # # 如果需要排除隐藏内容，传递参数给batch处理函数
        # if exclude_hidden:
        #     print("需要排除隐藏行和列，将在row_batch处理时执行")

        try:
            # 快速读取第一行做为列标题
            header_row = None
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                header_row = row
                break

            if not header_row:
                return None

            # 构建列定义
            columns = ['文件名']
            for idx, cell in enumerate(header_row):
                cell_value = cell.value
                if cell_value is not None:
                    columns.append(str(cell_value))
                else:
                    # 使用循环索引代替cell.column，避免EmptyCell错误
                    columns.append(f"列 {idx + 1}")

            # 添加额外列
            additional_columns = [f'Extra Column {i + 1}' for i in range(5)]
            columns.extend(additional_columns)

            file_basename = os.path.basename(file)
            found_rows = []

            # 预分配内存以提高性能
            all_rows = []
            batch_size = 100  # 每批处理的行数
            current_batch = []
            row_idx = 0

            # 按批次迭代行
            for row in worksheet.iter_rows(min_row=2):
                row_idx += 1
                current_batch.append(row)

                if len(current_batch) >= batch_size:
                    self._process_row_batch(
                        current_batch, row_idx - len(current_batch) + 1,
                        file, file_basename, sheet_name, keyword, match_mode,
                        columns, found_rows
                    )
                    current_batch = []

            # 处理剩余的行
            if current_batch:
                self._process_row_batch(
                    current_batch, row_idx - len(current_batch) + 1,
                    file, file_basename, sheet_name, keyword, match_mode,
                    columns, found_rows
                )

            # 如果没有找到匹配的结果，则返回None
            if not found_rows:
                return None

            # 返回结果
            return {
                'columns': columns,
                'rows': found_rows
            }

        except Exception as e:
            print(f"优化工作表搜索出错: {str(e)}")
            print(traceback.format_exc())
            return None

    def _process_row_batch(self, rows, start_idx, file, file_basename, sheet_name,
                           keyword, match_mode, columns, found_rows):
        """优化的批量处理行 - 提高处理速度和内存效率"""

        # 在调用_process_row_batch前添加
        if not hasattr(self, '_hidden_rows_cache'):
            self._hidden_rows_cache = {}
        if not hasattr(self, '_hidden_cols_cache'):
            self._hidden_cols_cache = {}
        if not hasattr(self, '_formula_cache'):
            self._formula_cache = {}

        # 一次性获取公式数据，减少重复加载
        formula_rows = None
        has_formula_data = False

        try:
            # 使用缓存键以避免重复加载相同工作簿
            cache_key = f"{file}:{sheet_name}"
            if cache_key in getattr(self, '_formula_cache', {}):
                formula_data = self._formula_cache[cache_key]
                if start_idx + 1 in formula_data and start_idx + len(rows) in formula_data:
                    # 缓存中已经存储的是单元格值的列表，不是单元格对象
                    formula_rows = [formula_data[start_idx + i + 1] for i in range(len(rows)) if
                                    start_idx + i + 1 in formula_data]
                    has_formula_data = True

            # 如果缓存未命中，加载工作簿
            if not has_formula_data:
                # 不使用with语句，改用传统的try-finally结构
                wb_formulas = None
                try:
                    wb_formulas = openpyxl.load_workbook(file, read_only=True, data_only=False)
                    ws_formulas = wb_formulas[sheet_name]
                    # 获取公式行数据
                    formula_rows_cells = list(
                        ws_formulas.iter_rows(min_row=start_idx + 1, max_row=start_idx + len(rows)))

                    # 从单元格对象中提取值，存储为列表
                    formula_rows = [[cell.value for cell in row] for row in formula_rows_cells]
                    has_formula_data = True

                    # 更新缓存
                    if not hasattr(self, '_formula_cache'):
                        self._formula_cache = {}
                    if cache_key not in self._formula_cache:
                        self._formula_cache[cache_key] = {}
                    for i, row in enumerate(formula_rows):
                        self._formula_cache[cache_key][start_idx + i + 1] = row
                finally:
                    # 确保无论如何都关闭工作簿
                    if wb_formulas:
                        try:
                            wb_formulas.close()
                        except:
                            pass

        except Exception as e:
            print(f"获取公式数据出错: {str(e)}")
            formula_rows = None
            has_formula_data = False

        # 使用批量处理的隐藏行列信息，减少Win32 COM调用
        hidden_rows = getattr(self, '_hidden_rows_cache', {}).get(f"{file}:{sheet_name}", set())
        hidden_cols = getattr(self, '_hidden_cols_cache', {}).get(f"{file}:{sheet_name}", set())
        exclude_hidden = getattr(self, 'exclude_hidden', None)
        exclude_hidden = exclude_hidden.get() if exclude_hidden else True

        # 在这里添加获取隐藏行信息的代码
        # 如果需要排除隐藏内容但缓存为空，则获取隐藏信息

        # 预处理关键字
        keyword_lower = keyword.lower() if isinstance(keyword, str) else str(keyword).lower()

        # 批量处理所有行
        matches = []
        for i, row in enumerate(rows):
            row_idx = start_idx + i
            excel_row_num = row_idx + 1

            # 快速检查是否隐藏行
            if exclude_hidden and excel_row_num in hidden_rows:
                continue

            # 获取当前行数据 - 使用列表推导式减少循环
            cell_values = [cell.value for cell in row]

            # 获取公式数据 - 避免不必要的处理
            if has_formula_data and i < len(formula_rows):
                # formula_rows现在是值的列表，不是单元格对象的列表
                formula_values = formula_rows[i]
            else:
                formula_values = []

            # 快速搜索文本 - 使用任何方式匹配
            found = False

            # 计算值搜索 - 合并为单一循环
            for idx, val in enumerate(cell_values):
                # 跳过隐藏列和空值
                col_idx = idx + 1
                if exclude_hidden and col_idx in hidden_cols:
                    continue

                if val is not None:
                    try:
                        # 直接处理字符串比较，减少转换
                        if isinstance(val, str):
                            cell_text = val.lower().strip()
                        else:
                            cell_text = str(val).lower().strip()

                        # 根据匹配模式比较
                        if (match_mode == "exact" and keyword_lower == cell_text) or \
                                (match_mode != "exact" and keyword_lower in cell_text):
                            found = True
                            break
                    except:
                        pass

            # 仅在计算值未匹配并且有公式值时搜索公式
            if not found and formula_values:
                for idx, val in enumerate(formula_values):
                    # 跳过隐藏列和空值
                    col_idx = idx + 1
                    if exclude_hidden and col_idx in hidden_cols:
                        continue

                    # 仅处理公式值
                    if val is not None and isinstance(val, str) and val.startswith('='):
                        try:
                            cell_text = val.lower().strip()

                            # 根据匹配模式比较
                            if (match_mode == "exact" and keyword_lower == cell_text) or \
                                    (match_mode != "exact" and keyword_lower in cell_text):
                                found = True
                                break
                        except:
                            pass

            # 如果找到匹配，构建结果
            if found:

                # 创建行值 - 一次性完成
                formula_row_values = [file_basename] + (formula_values if formula_values else [None] * len(cell_values))
                value_row_values = [file_basename] + cell_values

                # 确保列数一致 - 一次性操作
                expected_column_count = len(columns)
                if len(formula_row_values) < expected_column_count:
                    formula_row_values.extend([None] * (expected_column_count - len(formula_row_values)))
                if len(value_row_values) < expected_column_count:
                    value_row_values.extend([None] * (expected_column_count - len(value_row_values)))

                # 添加到临时结果列表
                matches.append({
                    'formula_values': formula_row_values,
                    'value_values': value_row_values,
                    'location': {
                        'file': file,
                        'sheet': sheet_name,
                        'row': row_idx,
                        'col_start': 1,
                        'col_end': len(cell_values) + 1
                    }
                })

        # 一次性添加所有匹配项到结果
        found_rows.extend(matches)

        # 手动触发垃圾回收以释放内存
        import gc
        if len(matches) > 100:  # 只在处理大量匹配时执行GC
            gc.collect()

    def save_changes(self):
        """优化版保存更改函数 - 使用并行处理和批量操作提升性能"""
        try:
            # 基本验证检查
            if self.filtered_data is None or not self.cell_locations:
                messagebox.showwarning("警告", "没有筛选后的数据可以保存")
                return

            modified_data = self.table.get_modified_data()
            if not modified_data:
                messagebox.showinfo("提示", "没有数据被修改")
                return

            # 调试输出
            print(f"Modified data: {modified_data}")
            print("== Row ID映射表 ==")
            for row_id, location in self.row_id_map.items():
                print(f"行ID: {row_id}, 位置: {location}")
            print("===============")

            # 创建进度窗口
            progress_window = tk.Toplevel(self)
            progress_window.title("保存更改")
            progress_window.geometry("400x150")
            progress_window.transient(self)
            progress_window.grab_set()

            progress_label = ttk.Label(progress_window, text="正在保存更改...")
            progress_label.pack(pady=(10, 5))
            progress_bar = ttk.Progressbar(progress_window, length=350, mode='determinate')
            progress_bar.pack(pady=5)
            detail_label = ttk.Label(progress_window, text="准备中...")
            detail_label.pack(pady=5)
            progress_window.update_idletasks()

            # 关键优化点1: 按文件分组修改，减少文件打开/关闭次数
            files_to_modify = {}  # {file_path: [(row_id, changes), ...]}
            row_id_to_location = {}  # 快速查找映射
            errors = []

            # 预处理 - 分组和验证
            for row_id, changes in modified_data.items():
                try:
                    # 确保row_id是公式行ID (以I或F开头)
                    if not (row_id.startswith('I') or row_id.startswith('F')):
                        if row_id.startswith('V'):
                            # 处理值行ID
                            idx = row_id[1:]  # 获取ID的数字部分
                            formula_id = f'I{idx}' if row_id[0] == 'V' else row_id

                            if formula_id not in self.row_id_map:
                                formula_id = f'F{idx}'  # 尝试另一种ID格式

                            if formula_id in self.row_id_map:
                                row_id = formula_id
                            else:
                                errors.append(f"找不到值行 '{row_id}' 对应的公式行")
                                continue
                        else:
                            errors.append(f"无效的行ID格式 '{row_id}'")
                            continue

                    if row_id not in self.row_id_map:
                        errors.append(f"找不到行ID对应的位置信息: {row_id}")
                        continue

                    location = self.row_id_map[row_id]
                    file = location.get('file')

                    # if not file or not os.path.exists(file):
                    #     errors.append(f"文件不存在: {file}")
                    #     continue

                    # 保存到文件分组字典
                    if file not in files_to_modify:
                        files_to_modify[file] = []

                    files_to_modify[file].append((row_id, changes))
                    row_id_to_location[row_id] = location

                except Exception as e:
                    error_msg = f"处理行 {row_id} 时出错: {str(e)}"
                    errors.append(error_msg)
                    print(error_msg)
                    print(traceback.format_exc())

            # 计算总任务数
            total_files = len(files_to_modify)
            if total_files == 0:
                progress_window.destroy()
                if errors:
                    error_msg = "\n".join(errors[:5])
                    if len(errors) > 5:
                        error_msg += f"\n...还有 {len(errors) - 5} 个错误未显示"
                    messagebox.showerror("保存失败", f"未能保存任何修改:\n{error_msg}")
                else:
                    messagebox.showinfo("提示", "没有修改被保存")
                return

            # 并行处理函数
            def process_file(file_args):
                file_path, modifications = file_args
                file_result = {
                    'file': file_path,
                    'success': False,
                    'modified_count': 0,
                    'errors': []
                }

                try:
                    # 打开工作簿 - 只打开一次
                    wb = openpyxl.load_workbook(file_path)

                    # 处理该文件中的所有修改
                    for row_id, changes in modifications:
                        try:
                            location = row_id_to_location[row_id]
                            sheet_name = location.get('sheet')
                            excel_row_num = location.get('row') + 1  # +1 因为在保存位置时行号可能做了-1处理
                            col_start = location.get('col_start', 1)
                            col_map = location.get('col_map', {})

                            # 确保工作表存在
                            if sheet_name not in wb.sheetnames:
                                file_result['errors'].append(f"工作表不存在: {sheet_name}")
                                continue

                            ws = wb[sheet_name]

                            # 检查行号是否有效
                            if excel_row_num < 1 or excel_row_num > ws.max_row:
                                file_result['errors'].append(
                                    f"无效的行号: {excel_row_num}, 工作表最大行数: {ws.max_row}")
                                continue

                            # 逐个处理单元格修改
                            for col_offset, value in changes.items():
                                # 确定Excel列号
                                if col_map and col_offset - 1 in col_map:
                                    col_num = col_map[col_offset - 1]
                                else:
                                    # 减去1来补偿文件名列的偏移
                                    adjusted_col_offset = col_offset - 1
                                    if adjusted_col_offset >= 0:  # 只处理非文件名列
                                        col_num = col_start + adjusted_col_offset
                                    else:
                                        continue  # 跳过文件名列

                                # 检查列号是否有效
                                if col_num < 1 or col_num > ws.max_column:
                                    file_result['errors'].append(
                                        f"无效的列号: {col_num}, 工作表最大列数: {ws.max_column}")
                                    continue

                                try:
                                    # 获取单元格并更新值
                                    cell = ws.cell(row=excel_row_num, column=col_num)
                                    # old_value = cell.value

                                    # 智能类型转换
                                    if isinstance(value, str):
                                        # 尝试转换为数值类型
                                        if value.strip() and all(c in '0123456789.-+' for c in value.strip()):
                                            try:
                                                if '.' in value:
                                                    value = float(value)
                                                else:
                                                    value = int(value)
                                            except ValueError:
                                                pass  # 保持为字符串

                                    # 更新单元格值
                                    cell.value = value
                                    file_result['modified_count'] += 1
                                except Exception as cell_err:
                                    file_result['errors'].append(
                                        f"更新单元格 ({excel_row_num}, {col_num}) 时出错: {str(cell_err)}")
                        except Exception as row_err:
                            file_result['errors'].append(f"处理行 {row_id} 时出错: {str(row_err)}")

                    # 全部处理完后一次性保存
                    wb.save(file_path)
                    file_result['success'] = True
                except PermissionError:
                    file_result['errors'].append(f"无法保存文件 '{file_path}': 文件可能被其他程序打开")
                except Exception as e:
                    file_result['errors'].append(f"保存文件 '{file_path}' 时出错: {str(e)}")
                finally:
                    # 确保工作簿已关闭
                    if 'wb' in locals():
                        try:
                            wb.close()
                            del wb
                        except:
                            pass

                return file_result

            # 关键优化点2: 使用线程池并行处理多个文件
            total_saved = 0
            file_items = list(files_to_modify.items())

            # 确定最佳线程数
            optimal_workers = min(os.cpu_count() or 4, len(file_items), 8)

            with concurrent.futures.ThreadPoolExecutor(max_workers=optimal_workers) as executor:
                # 提交所有任务
                futures = {executor.submit(process_file, file_item): file_item[0] for file_item in file_items}

                # 处理完成的任务
                for i, future in enumerate(concurrent.futures.as_completed(futures), 1):
                    file_path = futures[future]
                    progress = int(i * 100 / total_files)
                    progress_bar['value'] = progress
                    progress_label['text'] = f"正在保存更改... ({i}/{total_files})"
                    detail_label['text'] = f"处理文件: {os.path.basename(file_path)}"
                    progress_window.update_idletasks()

                    # 获取结果
                    result = future.result()
                    if result['success']:
                        total_saved += 1
                        errors.extend(result['errors'])
                    else:
                        errors.extend(result['errors'])

            # 处理完成，关闭进度窗口
            progress_window.destroy()

            # 显示结果
            if total_saved > 0:
                messagebox.showinfo("成功", f"成功保存了 {total_saved} 个文件的修改")


            else:
                if errors:
                    error_msg = "\n".join(errors[:5])
                    if len(errors) > 5:
                        error_msg += f"\n...还有 {len(errors) - 5} 个错误未显示"
                    messagebox.showerror("保存失败", f"未能保存任何修改:\n{error_msg}")
                else:
                    messagebox.showinfo("提示", "没有修改被保存")

            self.table.modified_cells.clear()

            try:
                print("保存完成，正在清理缓存并重新加载数据...")

                # 清理所有修改过的文件缓存
                for file_path in files_to_modify.keys():
                    # 处理全局缓存
                    if file_path in WORKBOOK_CACHE:
                        del WORKBOOK_CACHE[file_path]
                        print(f"已清除工作簿缓存: {os.path.basename(file_path)}")

                    if file_path in PARSED_DATA_CACHE:
                        del PARSED_DATA_CACHE[file_path]
                        print(f"已清除解析数据缓存: {os.path.basename(file_path)}")

                    # 处理实例缓存
                    for s_name in self.combined_data.keys():
                        cache_key = f"{file_path}:{s_name}"

                        if hasattr(self, '_formula_cache') and cache_key in self._formula_cache:
                            del self._formula_cache[cache_key]
                            print(f"已清除公式缓存: {os.path.basename(file_path)}:{s_name}")

                        if hasattr(self, '_hidden_rows_cache') and cache_key in self._hidden_rows_cache:
                            del self._hidden_rows_cache[cache_key]

                        if hasattr(self, '_hidden_cols_cache') and cache_key in self._hidden_cols_cache:
                            del self._hidden_cols_cache[cache_key]

                # 触发垃圾回收
                import gc
                gc.collect()

                # 强制重新加载数据
                if total_saved > 0:
                    print("文件已更改，重新加载数据...")
                    self.reload_data()  # 这会调用 load_all_data 和 filter_data

            except Exception as e:
                print(f"缓存清理或数据重载出错: {str(e)}")
                print(traceback.format_exc())

        except Exception as e:
            print(f"保存更改错误: {str(e)}")
            print(traceback.format_exc())
            messagebox.showerror("错误", f"保存更改时发生错误：{str(e)}")

    def get_table_data(self):
        data = []
        columns = self.filtered_data.columns if self.filtered_data is not None else []
        for item in self.table.get_children():
            values = self.table.item(item)['values']
            row_data = dict(zip(columns, values))
            data.append(row_data)
        return pd.DataFrame(data)

    def update_table(self, df):
        try:
            # 清除现有数据
            for item in self.table.get_children():
                self.table.delete(item)

            # 设置列
            num_columns = len(df.columns)
            self.table['columns'] = [f'#{i}' for i in range(num_columns)]

            # 设置列宽
            for i in range(num_columns):
                self.table.column(f'#{i}', width=100, stretch=tk.YES)

            # 创建行ID到位置信息的映射
            self.row_id_map = {}

            # 获取快速模式状态
            fast_mode = self.fast_mode.get()

            # 添加数据到表格，每两行为一组（公式行和值行）
            for i in range(0, len(df), 2):
                # 生成唯一的行ID
                row_id = f'I{i // 2:03d}'  # 使用数字格式而不是十六进制

                # 保存行ID到位置信息的映射
                if i // 2 < len(self.cell_locations):
                    self.row_id_map[row_id] = self.cell_locations[i // 2]
                    print(f"映射行ID {row_id} 到位置: {self.cell_locations[i // 2]}")

                # 插入公式行
                formula_values = list(df.iloc[i])
                formula_item = self.table.insert('', 'end', iid=row_id,
                                                 values=formula_values,
                                                 tags=('formula',))

                # 插入值行（如果存在）
                if i + 1 < len(df):
                    value_values = list(df.iloc[i + 1])

                    # 在快速模式下更简单地处理None值
                    if fast_mode:
                        for j in range(len(value_values)):
                            if value_values[j] is None and j < len(formula_values):
                                formula_val = formula_values[j]
                                if isinstance(formula_val, str) and formula_val.startswith('='):
                                    # 简单公式直接计算
                                    try:
                                        formula_expr = formula_val[1:].strip()
                                        if re.match(r'^[\d\s\+\-\*\/\(\)\.]+$', formula_expr):
                                            value_values[j] = eval(formula_expr)
                                        else:
                                            value_values[j] = ""  # 复杂公式显示为空
                                    except:
                                        value_values[j] = ""
                    else:
                        # 处理None值：如果计算值为None但公式存在，则显示实际值或"计算中..."
                        for j in range(len(value_values)):
                            if j < len(formula_values):
                                formula_val = formula_values[j]

                                # 如果计算值为None且公式行有公式，尝试直接计算
                                if value_values[j] is None and isinstance(formula_val, str) and formula_val.startswith(
                                        '='):
                                    try:
                                        # 尝试获取Excel公式对应的实际值
                                        value = None  # 默认为None

                                        # 除去公式前的"="，并尝试进行简单计算
                                        if '*' in formula_val or '+' in formula_val or '-' in formula_val or '/' in formula_val:
                                            # 有基本运算符，可能可以计算
                                            formula_expr = formula_val[1:].replace('*', '*').replace('/', '/')
                                            try:
                                                # 尝试安全地eval公式（仅限基本数学运算）
                                                value = eval(formula_expr)
                                            except:
                                                value = "计算中..."
                                        else:
                                            value = "计算中..."

                                        value_values[j] = value
                                    except:
                                        value_values[j] = "计算中..."

                    value_id = f'V{i // 2:03d}'
                    self.table.insert(formula_item, 'end', iid=value_id,
                                      values=value_values,
                                      tags=('value',))

            # 设置样式
            self.table.tag_configure('formula', background='#E6F3FF')
            self.table.tag_configure('value', background='#F0F8FF')

        except Exception as e:
            print(f"更新表格错误: {str(e)}")
            print(traceback.format_exc())

    def export_filtered_data(self):
        try:
            if not hasattr(self, 'filtered_data') or self.filtered_data is None:
                messagebox.showwarning("警告", "没有可导出的数据，请先进行搜索")
                return

            # 获取搜索关键词作为文件名
            keyword = self.keyword_input.get().strip()
            if not keyword:
                keyword = "筛选结果"  # 默文件名

            # 创建保存文件对话框
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"{keyword}.xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if not file_path:  # 用户取消了保存
                return

            # 创建新的Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = keyword

            row_out = 1  # 从第2行开始写数据
            for i in range(1, len(self.filtered_data), 2):  # 步长为2，只取值行
                if i < len(self.filtered_data):  # 确   有值行
                    row = self.filtered_data.iloc[i]  # 获取值行
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_out, column=col_idx, value=value)
                    row_out += 1
            # 保存文件
            wb.save(file_path)
            messagebox.showinfo("成功", f"数据已成功导出到：\n{file_path}")

        except Exception as e:
            print(f"导出数据错误: {str(e)}")
            print(traceback.format_exc())
            messagebox.showerror("错误", f"导出数据时出错：{str(e)}")

    def _process_excel_batch_win32(self, excel_app, files, sheet_name, keyword, match_mode, progress_callback=None):
        """使用Win32 COM API加速处理Excel批次 - 极致优化版"""
        print("\n=========== 正在使用 WIN32 COM 引擎处理 ===========\n")
        results = []

        try:
            # 设置Excel应用程序属性以最大化性能
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = False
            excel_app.EnableEvents = False
            excel_app.AskToUpdateLinks = False
            excel_app.Calculation = -4135  # xlCalculationManual

            # 使用额外优化设置
            try:
                # 禁用自动恢复
                excel_app.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
                excel_app.AutoRecover.Enabled = False
            except:
                pass

            # 获取总文件数以计算进度
            total_files = len(files)

            # 设置WIN32 COM API调用间隔 - 避免过多COM调用导致的内存泄漏
            batch_operations = 0
            COM_BATCH_THRESHOLD = 1000  # 每1000次COM操作后重置COM对象

            # 保存找到的列名
            found_columns = None

            # 预处理关键词以提高比较效率
            keyword_for_comparison = keyword.lower() if isinstance(keyword, str) else str(keyword).lower()

            # 对每个文件进行处理
            for file_idx, file in enumerate(files):
                if progress_callback:
                    progress_callback(int(file_idx * 100 / total_files), f"处理: {os.path.basename(file)}")

                try:
                    # 高级Excel打开参数设置
                    wb = excel_app.Workbooks.Open(
                        file,
                        ReadOnly=True,  # 只读模式
                        UpdateLinks=False,  # 不更新链接
                        IgnoreReadOnlyRecommended=True,  # 忽略只读建议
                        Notify=False,  # 不通知
                        AddToMru=False  # 不加入最近使用文件
                    )

                    try:
                        # 查找工作表
                        sheet = None
                        for ws in wb.Sheets:
                            batch_operations += 1
                            if ws.Name == sheet_name:
                                sheet = ws
                                break

                            # 定期检查COM操作计数并重置
                            if batch_operations >= COM_BATCH_THRESHOLD:
                                # 强制COM对象释放内存
                                import pythoncom
                                try:
                                    pythoncom.CoFreeUnusedLibraries()
                                except:
                                    pass
                                batch_operations = 0

                        if not sheet:
                            continue

                        # 获取使用范围
                        used_range = sheet.UsedRange
                        rows_count = used_range.Rows.Count
                        cols_count = used_range.Columns.Count

                        # 一次性获取所有数据，减少COM调用
                        data_variant = used_range.Value
                        formulas_variant = used_range.Formula
                        hidden_rows_collection = []
                        hidden_cols_collection = []

                        batch_operations += 3  # 增加COM操作计数

                        # 处理数据和公式数组 - 高效转换
                        data_array = self._convert_com_data(data_variant, rows_count, cols_count)
                        formula_array = self._convert_com_data(formulas_variant, rows_count, cols_count)

                        # 批量获取隐藏行列信息 - 减少COM调用
                        try:
                            # 使用SpecialCells方法一次性获取所有隐藏行
                            try:
                                hidden_ranges = sheet.Cells.SpecialCells(12).Areas  # xlCellTypeVisible的补集
                                for area in hidden_ranges:
                                    hidden_row_range = area.Rows
                                    for i in range(1, hidden_row_range.Count + 1):
                                        hidden_rows_collection.append(hidden_row_range(i).Row)
                            except:
                                # 如果SpecialCells方法失败，退回到逐行检查
                                # 但使用更高效的步进方式，不是逐行查询
                                row_check_step = max(1, rows_count // 20)  # 每次检查多行
                                for r in range(1, rows_count + 1, row_check_step):
                                    end_r = min(r + row_check_step - 1, rows_count)
                                    row_range = sheet.Rows(f"{r}:{end_r}")
                                    if row_range.Hidden:
                                        for i in range(r, end_r + 1):
                                            hidden_rows_collection.append(i)
                                    batch_operations += 1

                            # 同样高效获取隐藏列
                            col_check_step = max(1, cols_count // 20)
                            for c in range(1, cols_count + 1, col_check_step):
                                end_c = min(c + col_check_step - 1, cols_count)
                                col_letter_start = openpyxl.utils.get_column_letter(c)
                                col_letter_end = openpyxl.utils.get_column_letter(end_c)
                                col_range = sheet.Columns(f"{col_letter_start}:{col_letter_end}")
                                if col_range.Hidden:
                                    for i in range(c, end_c + 1):
                                        hidden_cols_collection.append(i)
                                batch_operations += 1
                        except:
                            # 如果批量获取隐藏行列失败，则使用空列表
                            print(f"批量获取隐藏行列信息失败: {file}")

                        # 转换为集合以提高查询效率
                        hidden_rows = set(hidden_rows_collection)
                        hidden_cols = set(hidden_cols_collection)

                        # 定义列名 - 优化处理
                        if rows_count > 0 and not found_columns:
                            temp_columns = ['文件名']

                            # 从第一行提取列名
                            if len(data_array) > 0:
                                header_row = data_array[0]
                                for c in range(len(header_row)):
                                    col_idx = c + 1  # 列索引从1开始
                                    if col_idx not in hidden_cols:
                                        cell_value = header_row[c]
                                        if cell_value is not None:
                                            temp_columns.append(cell_value)

                            # 添加额外列
                            additional_columns = [f'Extra Column {i + 1}' for i in range(5)]
                            temp_columns.extend(additional_columns)

                            # 保存找到的列名
                            found_columns = temp_columns

                        # 使用预设的列名或默认列名
                        columns = found_columns if found_columns else ['文件名'] + [f'列 {i}' for i in
                                                                                    range(1, cols_count + 6)]

                        # 查找匹配行 - 使用预处理的数据数组，减少COM调用
                        found_rows = []
                        file_basename = os.path.basename(file)

                        # 使用批处理模式处理行 - 基于数组直接处理，避免COM调用
                        for r in range(1, rows_count):  # 跳过标题行
                            if r + 1 in hidden_rows:  # r+1因为数组索引从0开始，而行索引从1开始
                                continue

                            # 获取当前行的公式和值
                            formula_row = formula_array[r] if r < len(formula_array) else []
                            value_row = data_array[r] if r < len(data_array) else []

                            # 快速搜索标志
                            found = False

                            # 检查整行是否包含关键词 - 使用高效的字符串处理
                            row_text_list = []
                            row_text = ""

                            # 构建搜索文本
                            for c in range(len(value_row)):
                                col_idx = c + 1
                                if col_idx not in hidden_cols:
                                    value = value_row[c]
                                    if value is not None:
                                        # 添加处理后的文本到列表
                                        try:
                                            text = str(value).lower().strip()
                                            row_text_list.append(text)
                                            row_text += " " + text
                                        except:
                                            pass

                            # 根据匹配模式检查是否找到关键词
                            if match_mode == "exact":
                                # 精确匹配 - 检查列表中是否有完全匹配的项
                                for text in row_text_list:
                                    if text == keyword_for_comparison:
                                        found = True
                                        break
                            else:
                                # 模糊匹配 - 检查整行文本是否包含关键词
                                found = keyword_for_comparison in row_text

                            # 如果找到匹配，构建完整的行数据
                            if found:
                                # 预分配行数据数组
                                formula_values = [file_basename]
                                value_values = [file_basename]

                                # 填充formula_values和value_values数组
                                for c in range(len(formula_row)):
                                    formula_values.append(formula_row[c])

                                for c in range(len(value_row)):
                                    value_values.append(value_row[c])

                                # 确保列数一致
                                expected_column_count = len(columns)
                                if len(formula_values) < expected_column_count:
                                    formula_values.extend([None] * (expected_column_count - len(formula_values)))
                                if len(value_values) < expected_column_count:
                                    value_values.extend([None] * (expected_column_count - len(value_values)))

                                # 添加到结果
                                found_rows.append({
                                    'formula_values': formula_values,
                                    'value_values': value_values,
                                    'location': {
                                        'file': file,
                                        'sheet': sheet_name,
                                        'row': r + 1,  # 行号从1开始
                                        'col_start': 1,
                                        'col_end': cols_count
                                    }
                                })

                        # 添加文件结果
                        results.append({
                            'columns': columns,
                            'rows': found_rows
                        })

                    finally:
                        # 关闭工作簿
                        wb.Close(SaveChanges=False)
                        batch_operations += 1

                except Exception as e:
                    print(f"使用Win32处理文件 '{file}' 时出错: {str(e)}")
                    results.append(None)

                # 定期释放COM对象内存
                if batch_operations >= COM_BATCH_THRESHOLD:
                    try:
                        import pythoncom
                        pythoncom.CoFreeUnusedLibraries()
                    except:
                        pass
                    batch_operations = 0

            # 恢复Excel设置
            excel_app.Calculation = -4105  # xlCalculationAutomatic
            excel_app.ScreenUpdating = True
            excel_app.EnableEvents = True

        except Exception as e:
            print(f"Win32批处理错误: {str(e)}")
            print(traceback.format_exc())

        return results

    def _convert_com_data(self, com_variant, rows_count, cols_count):
        """高效转换COM变体数据到Python数组"""
        try:
            # 首先处理空值情况
            if com_variant is None:
                return []

            # 处理基本类型（非数组）
            if not isinstance(com_variant, tuple):
                return [[com_variant]]

            # 处理一维数组
            if rows_count == 1:
                if isinstance(com_variant, tuple) and len(com_variant) > 0:
                    # 确保是一个二维数组
                    return [com_variant]
                else:
                    return [tuple([com_variant])]

            # 处理二维数组 - 直接返回
            return com_variant
        except Exception as e:
            print(f"转换COM数据出错: {str(e)}")
            # 返回空数组作为后备
            return []

    def process_excel_batch(self, batch_files):
        """处理一批Excel文件的公式预计算并自动处理安全警告"""
        # 初始化COM环境
        pythoncom.CoInitialize()

        processed = []
        errors = []
        excel = None

        try:
            # 创建Excel应用实例
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False  # 尝试禁用一些警告
            excel.EnableEvents = False
            excel.ScreenUpdating = False
            excel.AskToUpdateLinks = False

            # 处理批次中的每个文件
            for file_path in batch_files:
                try:
                    # 打开工作簿
                    wb = excel.Workbooks.Open(file_path, UpdateLinks=False, ReadOnly=False)

                    # 提前收集所有工作表，避免在循环中重复访问COM对象
                    sheets = []
                    try:
                        for i in range(1, wb.Sheets.Count + 1):
                            sheet = wb.Sheets(i)
                            if sheet.Type == -4167:  # xlWorksheet
                                sheets.append(sheet)
                    except Exception as e:
                        print(f"收集工作表时出错: {str(e)}")

                    try:
                        # 单独计算每个工作表
                        for sheet in sheets:
                            try:
                                sheet.Calculate()
                            except Exception as se:
                                print(f"计算工作表时出错: {str(se)}")
                    except Exception as e:
                        print(f"遍历工作表时出错: {str(e)}")

                    # 确保所有外部数据已刷新
                    try:
                        if wb.Connections.Count > 0:
                            wb.RefreshAll()
                    except Exception as e:
                        print(f"刷新连接时出错: {str(e)}")

                    # 保存并关闭
                    wb.Save()
                    wb.Close(True)

                    processed.append(file_path)

                except Exception as e:
                    errors.append((file_path, str(e)))
                    print(f"处理文件 '{file_path}' 时出错: {str(e)}")

        finally:
            # 关闭Excel应用
            if excel:
                try:
                    excel.Quit()
                except:
                    pass

                # 强制释放COM对象
                del excel
                gc.collect()

            # 取消COM初始化
            pythoncom.CoUninitialize()

        return processed, errors

    def precalculate_excel_files(self):
        """批量预计算所有已加载的Excel文件的公式"""
        if not self.excel_files:
            messagebox.showinfo("提示", "请先加载Excel文件")
            return

        # 创建进度窗口
        progress_window = tk.Toplevel(self)
        progress_window.title("公式预计算")
        progress_window.geometry("400x200")
        progress_window.transient(self)
        progress_window.grab_set()

        # 进度显示组件
        progress_label = ttk.Label(progress_window, text="正在预计算Excel公式...")
        progress_label.pack(pady=(20, 10))

        progress_bar = ttk.Progressbar(progress_window, length=350, mode='determinate')
        progress_bar.pack(pady=10)

        detail_label = ttk.Label(progress_window, text="准备中...")
        detail_label.pack(pady=10)

        stats_frame = ttk.Frame(progress_window)
        stats_frame.pack(pady=5, fill='x', padx=20)

        ttk.Label(stats_frame, text="已处理:").grid(row=0, column=0, sticky='w')
        files_processed_label = ttk.Label(stats_frame, text="0/0")
        files_processed_label.grid(row=0, column=1, sticky='w', padx=5)

        # 确定最佳并行数
        max_workers = min(os.cpu_count() or 4, 4)  # 最多4个Excel实例

        # 优化批处理分配策略
        if len(self.excel_files) <= max_workers:
            # 文件数少于或等于工作线程数，每个线程处理一个文件
            batches = [[file] for file in self.excel_files]
        else:
            # 平均分配文件到每个线程
            batch_size = len(self.excel_files) // max_workers
            remainder = len(self.excel_files) % max_workers

            batches = []
            start_idx = 0
            for i in range(max_workers):
                # 为前remainder个批次多分配一个文件
                end_idx = start_idx + batch_size + (1 if i < remainder else 0)
                batches.append(self.excel_files[start_idx:end_idx])
                start_idx = end_idx

        def update_progress(current, total, status_text=""):
            progress = int(current / total * 100)
            self.after(0, lambda: progress_bar.configure(value=progress))
            self.after(0, lambda: files_processed_label.configure(text=f"{current}/{total}"))
            if status_text:
                self.after(0, lambda: detail_label.configure(text=status_text))

        def process_all_batches():
            start_time = time.time()
            processed_count = 0
            error_count = 0

            update_progress(0, len(self.excel_files), "开始处理...")

            # 使用线程池并行处理
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = []
                for batch in batches:
                    if batch:  # 确保批次不为空
                        futures.append(executor.submit(self.process_excel_batch, batch))

                # 等待所有任务完成
                for future in concurrent.futures.as_completed(futures):
                    try:
                        batch_processed, batch_errors = future.result()
                        processed_count += len(batch_processed)
                        error_count += len(batch_errors)

                        # 更新进度
                        update_progress(processed_count, len(self.excel_files),
                                        f"已处理: {processed_count}/{len(self.excel_files)} 文件")

                    except Exception as e:
                        error_msg = str(e)
                        print(f"批处理出错: {error_msg}")
                        self.after(0, lambda e=error_msg: messagebox.showerror("错误", f"处理失败: {e}"))

            elapsed = time.time() - start_time
            self.after(0, lambda: progress_window.destroy())

            result_message = f"已处理 {processed_count} 个文件，用时: {elapsed:.2f}秒"
            if error_count > 0:
                result_message += f"\n处理过程中发生 {error_count} 个错误"

            self.after(0, lambda msg=result_message: messagebox.showinfo("完成", msg))

        # 启动处理线程
        threading.Thread(target=process_all_batches, daemon=True).start()

    # 规则审计相关方法
    def _check_required_libraries(self):
        """检查是否安装了所有必要的库"""
        try:
            # 检查xlrd库（用于读取.xls文件）
            import importlib
            xlrd_spec = importlib.util.find_spec("xlrd")
            if xlrd_spec is None:
                messagebox.showwarning(
                    "缺少依赖库",
                    "未安装xlrd库，将无法读取旧版Excel（.xls）格式的规则文件。\n\n"
                    "建议通过以下命令安装xlrd库：\n"
                    "pip install xlrd==1.2.0\n\n"
                    "注意: 较新版本的xlrd可能不支持.xls格式，建议使用1.2.0版本。\n\n"
                    "如果不需要处理.xls文件，可以忽略此警告。"
                )
                return False
            return True
        except Exception as e:
            print(f"检查依赖库时出错: {str(e)}")
            return False
            
    def audit_rules(self):
        """打开规则文件并执行规则审计"""
        try:
            # 检查必要的库
            self._check_required_libraries()
            
            if not self.excel_files:
                messagebox.showwarning("警告", "请先加载Excel文件")
                return

            # 选择规则文件
            rule_file = filedialog.askopenfilename(
                title="选择规则文件",
                filetypes=[("Excel文件", "*.xlsx;*.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
            )

            if not rule_file:
                return  # 用户取消了选择

            # 创建进度窗口
            progress_window = tk.Toplevel(self)
            progress_window.title("规则审计")
            progress_window.geometry("400x150")
            progress_window.transient(self)
            progress_window.grab_set()

            progress_label = ttk.Label(progress_window, text="正在执行规则审计...")
            progress_label.pack(pady=(10, 5))
            progress_bar = ttk.Progressbar(progress_window, length=350, mode='determinate')
            progress_bar.pack(pady=5)
            detail_label = ttk.Label(progress_window, text="加载规则文件...")
            detail_label.pack(pady=5)
            progress_window.update_idletasks()

            # 解析规则文件
            parser = RuleParser(rule_file)
            rules = parser.parse_rules()

            if not rules:
                progress_window.destroy()
                messagebox.showwarning("警告", "规则文件为空或格式不正确")
                return

            # 执行规则验证
            validator = RuleValidator(self.excel_files)
            progress_bar['value'] = 10
            detail_label['text'] = "验证规则中..."
            progress_window.update_idletasks()

            # 启动验证过程
            def run_validation():
                try:
                    results = validator.validate_rules(rules, 
                        lambda p, m: self.after(0, lambda: self._update_audit_progress(progress_bar, detail_label, p, m)))
                    self.after(0, lambda: self._show_audit_report(results, progress_window))
                except Exception as e:
                    self.after(0, lambda: progress_window.destroy())
                    self.after(0, lambda: messagebox.showerror("错误", f"审计过程中出错：{str(e)}"))
                    print(f"审计错误: {str(e)}")
                    print(traceback.format_exc())

            threading.Thread(target=run_validation, daemon=True).start()

        except Exception as e:
            messagebox.showerror("错误", f"启动审计过程失败: {str(e)}")
            print(f"审计启动错误: {str(e)}")
            print(traceback.format_exc())

    def _update_audit_progress(self, progress_bar, detail_label, progress, message):
        """更新审计进度"""
        progress_bar['value'] = progress
        detail_label['text'] = message
        
    def _show_audit_report(self, results, progress_window):
        """显示审计报告"""
        progress_window.destroy()
        
        if not results or not results.get('errors'):
            messagebox.showinfo("审计完成", "未发现错误，所有规则验证通过！")
            return
            
        # 创建报告窗口
        report_window = tk.Toplevel(self)
        report_window.title("审计报告")
        report_window.geometry("800x600")
        
        # 创建报告显示区域
        frame = ttk.Frame(report_window, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 添加标题
        ttk.Label(frame, text="Excel审计报告", font=("Arial", 16, "bold")).pack(pady=(0, 10))
        
        # 添加摘要信息
        summary_frame = ttk.Frame(frame)
        summary_frame.pack(fill=tk.X, pady=(0, 10))
        
        total_rules = results.get('total_rules', 0)
        errors_count = len(results.get('errors', []))
        pass_rate = 1 - (errors_count / total_rules) if total_rules > 0 else 0
        
        ttk.Label(summary_frame, text=f"总规则数: {total_rules}").pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(summary_frame, text=f"错误数: {errors_count}").pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(summary_frame, text=f"通过率: {pass_rate:.2%}").pack(side=tk.LEFT)
        
        # 创建一个专门的 frame 用于 Treeview 和滚动条
        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # 创建错误列表
        columns = ("文件", "子表", "单元格", "当前值", "期望值", "规则类型")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", style='Custom.Treeview')
        
        # 设置列标题和宽度
        tree.heading("文件", text="文件")
        tree.column("文件", width=150)
        
        tree.heading("子表", text="子表")
        tree.column("子表", width=80)
        
        tree.heading("单元格", text="单元格")
        tree.column("单元格", width=80)
        
        tree.heading("当前值", text="当前值")
        tree.column("当前值", width=200)
        
        tree.heading("期望值", text="期望值")
        tree.column("期望值", width=200)
        
        tree.heading("规则类型", text="规则类型")
        tree.column("规则类型", width=100)
        
        # 添加滚动条
        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # 在 tree_frame 中使用 grid 布局
        tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        
        # 配置 tree_frame 的网格权重
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 填充数据
        for error in results.get('errors', []):
            tree.insert("", tk.END, values=(
                os.path.basename(error.get('file', '')),
                error.get('sheet', ''),
                error.get('cell', ''),
                error.get('current_value', ''),
                error.get('expected_value', ''),
                error.get('rule_type', '')
            ))
        
        # 添加导出按钮
        button_frame = ttk.Frame(report_window, padding=10)
        button_frame.pack(fill=tk.X)
        
        def export_report():
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            if file_path:
                reporter = RuleReporter()
                reporter.export_report(results, file_path)
                messagebox.showinfo("导出成功", f"报告已导出至：{file_path}")
        
        ttk.Button(button_frame, text="导出报告", command=export_report).pack(side=tk.RIGHT)


class RuleReporter:
    """规则审计报告生成器"""
    
    def export_report(self, results, file_path):
        """导出审计报告到Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # 创建新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "审计报告"
            
            # 设置标题样式
            title_font = Font(name='Arial', size=14, bold=True)
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 添加标题
            ws['A1'] = "Excel审计报告"
            ws.merge_cells('A1:F1')
            ws['A1'].font = title_font
            ws['A1'].alignment = center_alignment
            
            # 添加生成时间
            ws['A2'] = f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws.merge_cells('A2:F2')
            
            # 添加表头
            headers = ["文件", "子表", "单元格", "当前值", "期望值", "规则类型"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
                
            # 调整列宽
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 20
            
            # 添加数据
            row_idx = 5
            for error in results.get('errors', []):
                ws.cell(row=row_idx, column=1).value = os.path.basename(error.get('file', ''))
                ws.cell(row=row_idx, column=2).value = error.get('sheet', '')
                ws.cell(row=row_idx, column=3).value = error.get('cell', '')
                ws.cell(row=row_idx, column=4).value = error.get('current_value', '')
                ws.cell(row=row_idx, column=5).value = error.get('expected_value', '')
                ws.cell(row=row_idx, column=6).value = error.get('rule_type', '')
                
                # 设置单元格边框
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).border = border
                    
                row_idx += 1
                
            # 添加摘要信息
            ws.cell(row=row_idx + 2, column=1).value = "审计摘要"
            ws.merge_cells(f'A{row_idx + 2}:F{row_idx + 2}')
            ws.cell(row=row_idx + 2, column=1).font = Font(name='Arial', size=12, bold=True)
            
            ws.cell(row=row_idx + 3, column=1).value = "总规则数:"
            ws.cell(row=row_idx + 3, column=2).value = results.get('total_rules', 0)
            
            ws.cell(row=row_idx + 4, column=1).value = "错误数:"
            ws.cell(row=row_idx + 4, column=2).value = len(results.get('errors', []))
            
            ws.cell(row=row_idx + 5, column=1).value = "通过率:"
            total_rules = results.get('total_rules', 0)
            if total_rules > 0:
                pass_rate = 1 - (len(results.get('errors', [])) / total_rules)
                ws.cell(row=row_idx + 5, column=2).value = f"{pass_rate:.2%}"
            else:
                ws.cell(row=row_idx + 5, column=2).value = "N/A"
                
            # 保存文件
            wb.save(file_path)
            
        except Exception as e:
            print(f"导出审计报告时出错: {str(e)}")
            print(traceback.format_exc())
            raise


class RuleValidator:
    """规则验证器"""
    
    def __init__(self, excel_files):
        self.excel_files = excel_files
        self.errors = []
        
    def _is_valid_data_cell(self, cell):
        """判断单元格是否包含有效数据（非表头文本）
        
        有效数据定义：
        1. 数值类型（int, float）
        2. 以'='开头的公式字符串
        3. 公式计算后的结果
        
        表头定义：
        不以'='开头的纯文本字符串
        """
        if cell.value is None or cell.value == "":
            return False
            
        # 如果是数值类型，则为有效数据
        if isinstance(cell.value, (int, float)):
            return True
            
        # 如果是字符串且以'='开头，则为公式
        if isinstance(cell.value, str):
            if cell.value.startswith('='):
                return True
            # 否则，判断是否为数值字符串
            try:
                float(cell.value)
                return True  # 可以转换为数值
            except ValueError:
                return False  # 不可转换为数值，可能是表头文本
                
        # 对于其他类型，保持谨慎，返回True
        return True

    def _is_pure_aggregate_formula(self, formula, agg_func, target_col):
        """检查公式是否为纯聚合函数（不包含额外运算）

        Args:
            formula: 公式字符串，如 "=SUM(R9:R22)"
            agg_func: 聚合函数名，如 "SUM"
            target_col: 目标列，如 "R"

        Returns:
            bool: 如果是纯聚合函数返回True，否则返回False
        """
        if not formula or not formula.startswith('='):
            return False

        # 移除等号和空格
        formula_content = formula[1:].strip()

        # 构建纯聚合函数的正则表达式模式
        # 匹配格式：FUNC(COLx:COLy) 或 FUNC(COLx:COLy,...)
        # 例如：SUM(R9:R22) 或 SUM(R9:R22,R25:R30) 或 SUM( R9 : R22 )
        pattern = rf'^{agg_func}\s*\(\s*{target_col}\s*\d+\s*:\s*{target_col}\s*\d+(?:\s*,\s*{target_col}\s*\d+\s*:\s*{target_col}\s*\d+)*\s*\)$'

        # 使用正则表达式进行严格匹配
        return bool(re.match(pattern, formula_content, re.IGNORECASE))

    def validate_rules(self, rules, progress_callback=None):
        """验证所有规则"""
        total_rules = len(rules)
        self.errors = []
        
        # 预处理规则，按文件和工作表分组
        grouped_rules = self._group_rules_by_sheet(rules)
        
        # 处理每个文件
        for file_idx, file_path in enumerate(self.excel_files):
            file_basename = os.path.basename(file_path)
            if progress_callback:
                progress_callback(10 + (file_idx * 80 / len(self.excel_files)), 
                                 f"处理文件 {file_idx+1}/{len(self.excel_files)}: {file_basename}")
            
            try:
                # 打开工作簿
                wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
                
                # 获取工作表名称
                sheet_names = wb.sheetnames
                
                # 处理每个有规则的工作表
                for sheet_name, sheet_rules in grouped_rules.items():
                    if sheet_name not in sheet_names:
                        continue
                        
                    # 获取工作表
                    ws = wb[sheet_name]
                    
                    # 处理该工作表的所有规则
                    for rule_idx, rule in enumerate(sheet_rules):
                        if progress_callback:
                            rule_progress = ((file_idx * len(grouped_rules) + list(grouped_rules.keys()).index(sheet_name)) * 100 / 
                                           (len(self.excel_files) * len(grouped_rules)))
                            progress_callback(10 + rule_progress * 0.8, 
                                             f"验证规则 {rule_idx+1}/{len(sheet_rules)}: {sheet_name}.{rule['cell_ref']}")
                        
                        # 根据规则类型执行验证
                        rule_type = rule.get('rule_type', 'FIXED_VALUE')
                        
                        if rule_type == 'FIXED_VALUE':
                            self._validate_fixed_value(wb, ws, rule, file_path)
                        elif rule_type == 'DYN_CELL_PARTIAL_MATCH':
                            self._validate_dyn_cell_partial_match(wb, ws, rule, file_path)
                        elif rule_type == 'COL_FORMULA':
                            self._validate_col_formula(wb, ws, rule, file_path)
                        elif rule_type == 'COL_AGGREGATE_CHECK':
                            self._validate_col_aggregate_check(wb, ws, rule, file_path)
                        else:
                            print(f"不支持的规则类型: {rule_type}")
                
                # 关闭工作簿
                wb.close()
                
            except Exception as e:
                self.errors.append({
                    'file': file_path,
                    'sheet': 'N/A',
                    'cell': 'N/A',
                    'current_value': 'N/A',
                    'expected_value': 'N/A',
                    'rule_type': 'ERROR',
                    'error_message': str(e)
                })
                print(f"验证文件 '{file_path}' 时出错: {str(e)}")
                print(traceback.format_exc())
        
        if progress_callback:
            progress_callback(100, "验证完成")
            
        return {
            'errors': self.errors,
            'total_rules': total_rules
        }
        
    def _group_rules_by_sheet(self, rules):
        """将规则按工作表分组"""
        grouped = {}
        for rule in rules:
            sheet_name = rule.get('sheet', '')
            if sheet_name not in grouped:
                grouped[sheet_name] = []
            grouped[sheet_name].append(rule)
        return grouped
        
    def _validate_fixed_value(self, wb, ws, rule, file_path):
        """验证固定值规则"""
        cell_ref = rule.get('cell_ref', '')
        expected_value = rule.get('expected_value', '')
        
        # 如果cell_ref指定的是单个单元格
        if re.match(r'^[A-Za-z]+\d+$', cell_ref):
            try:
                # 获取单元格的值和公式
                cell = ws[cell_ref]
                current_formula = cell.value if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') else None
                
                # 如果有公式，获取公式字符串
                if current_formula is None and hasattr(cell, 'formula') and cell.formula:
                    current_formula = f"={cell.formula}" if not cell.formula.startswith('=') else cell.formula
                
                # 如果有公式，使用公式进行比较，否则使用值
                current_value = current_formula if current_formula else cell.value
                
                # 将值转换为字符串进行比较
                str_current = str(current_value) if current_value is not None else ""
                str_expected = str(expected_value) if expected_value is not None else ""
                
                # 检查是否匹配
                if str_current != str_expected:
                    self.errors.append({
                        'file': file_path,
                        'sheet': rule.get('sheet', ''),
                        'cell': cell_ref,
                        'current_value': str_current,
                        'expected_value': str_expected,
                        'rule_type': 'FIXED_VALUE'
                    })
            except Exception as e:
                self.errors.append({
                    'file': file_path,
                    'sheet': rule.get('sheet', ''),
                    'cell': cell_ref,
                    'current_value': 'ERROR',
                    'expected_value': expected_value,
                    'rule_type': 'FIXED_VALUE',
                    'error_message': str(e)
                })
                print(f"验证单元格 '{cell_ref}' 时出错: {str(e)}")
        
        # 如果cell_ref指定的是一个列（如"A列"）
        elif re.match(r'^[A-Za-z]+列$', cell_ref) or re.match(r'^[A-Za-z]+$', cell_ref):
            # 提取列字母
            col_letter = re.match(r'^([A-Za-z]+)', cell_ref).group(1)
            self._validate_column(wb, ws, col_letter, expected_value, rule, file_path)
            
    def _validate_column(self, wb, ws, col_letter, expected_value, rule, file_path):
        """验证列的值或公式"""
        # 获取工作表的使用范围
        min_row = ws.min_row
        max_row = ws.max_row
        
        # 遍历该列的所有单元格
        for row in range(min_row, max_row + 1):
            cell_ref = f"{col_letter}{row}"
            try:
                cell = ws[cell_ref]
                
                # 获取单元格的公式或值
                current_formula = cell.value if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') else None
                
                if current_formula is None and hasattr(cell, 'formula') and cell.formula:
                    current_formula = f"={cell.formula}" if not cell.formula.startswith('=') else cell.formula
                
                current_value = current_formula if current_formula else cell.value
                
                # 如果期望值以"="开头，则假定是公式验证
                if expected_value.startswith('='):
                    str_current = str(current_formula) if current_formula is not None else ""
                    if str_current != expected_value:
                        self.errors.append({
                            'file': file_path,
                            'sheet': rule.get('sheet', ''),
                            'cell': cell_ref,
                            'current_value': str_current,
                            'expected_value': expected_value,
                            'rule_type': 'COLUMN_FORMULA'
                        })
            except Exception as e:
                print(f"验证单元格 '{cell_ref}' 时出错: {str(e)}")
                
    def _validate_dyn_cell_partial_match(self, wb, ws, rule, file_path):
        """验证动态单元格部分匹配规则"""
        cell_ref = rule.get('cell_ref', '')
        tpl = rule.get('TPL', '')
        
        if not tpl or '{cell}' not in tpl:
            return
            
        # 分解模板为前缀和后缀
        prefix, suffix = tpl.split('{cell}', 1)
        
        # 如果cell_ref指定的是单个单元格
        if re.match(r'^[A-Za-z]+\d+$', cell_ref):
            try:
                # 获取单元格的公式
                cell = ws[cell_ref]
                current_formula = cell.value if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') else None
                
                if current_formula is None and hasattr(cell, 'formula') and cell.formula:
                    current_formula = f"={cell.formula}" if not cell.formula.startswith('=') else cell.formula
                
                # 如果没有公式，跳过
                if not current_formula:
                    return
                
                # 检查公式是否以前缀开头并以后缀结尾
                if not (current_formula.startswith(prefix) and current_formula.endswith(suffix)):
                    self.errors.append({
                        'file': file_path,
                        'sheet': rule.get('sheet', ''),
                        'cell': cell_ref,
                        'current_value': current_formula,
                        'expected_value': tpl.replace('{cell}', '*'),
                        'rule_type': 'DYN_CELL_PARTIAL_MATCH'
                    })
            except Exception as e:
                self.errors.append({
                    'file': file_path,
                    'sheet': rule.get('sheet', ''),
                    'cell': cell_ref,
                    'current_value': 'ERROR',
                    'expected_value': tpl,
                    'rule_type': 'DYN_CELL_PARTIAL_MATCH',
                    'error_message': str(e)
                })
                print(f"验证动态单元格 '{cell_ref}' 时出错: {str(e)}")
        
        # 如果cell_ref指定的是一个列（如"A列"）
        elif re.match(r'^[A-Za-z]+列$', cell_ref) or re.match(r'^[A-Za-z]+$', cell_ref):
            # 提取列字母
            col_letter = re.match(r'^([A-Za-z]+)', cell_ref).group(1)
            
            # 获取工作表的使用范围
            min_row = ws.min_row
            max_row = ws.max_row
            
            # 遍历该列的所有单元格
            for row in range(min_row, max_row + 1):
                cell_ref = f"{col_letter}{row}"
                try:
                    cell = ws[cell_ref]
                    
                    # 获取单元格的公式
                    current_formula = cell.value if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') else None
                    
                    if current_formula is None and hasattr(cell, 'formula') and cell.formula:
                        current_formula = f"={cell.formula}" if not cell.formula.startswith('=') else cell.formula
                    
                    # 如果没有公式，跳过
                    if not current_formula:
                        continue
                    
                    # 检查公式是否以前缀开头并以后缀结尾
                    if not (current_formula.startswith(prefix) and current_formula.endswith(suffix)):
                        self.errors.append({
                            'file': file_path,
                            'sheet': rule.get('sheet', ''),
                            'cell': cell_ref,
                            'current_value': current_formula,
                            'expected_value': tpl.replace('{cell}', '*'),
                            'rule_type': 'DYN_CELL_PARTIAL_MATCH'
                        })
                except Exception as e:
                    print(f"验证动态单元格 '{cell_ref}' 时出错: {str(e)}")
                    
    def _validate_col_formula(self, wb, ws, rule, file_path):
        """验证整列公式规则"""
        cell_ref = rule.get('cell_ref', '')
        tpl = rule.get('TPL', '')
        cond_col = rule.get('COND_COL', '')
        
        if not tpl or not cell_ref:
            return
            
        # 获取需要验证的列字母
        col_letter = None
        if re.match(r'^[A-Za-z]+列$', cell_ref):
            col_letter = re.match(r'^([A-Za-z]+)', cell_ref).group(1)
        elif re.match(r'^[A-Za-z]+$', cell_ref):
            col_letter = cell_ref
            
        if not col_letter:
            return
            
        # 获取条件列字母
        cond_col_letter = cond_col if re.match(r'^[A-Za-z]+$', cond_col) else None
        
        # 获取工作表的使用范围
        min_row = ws.min_row
        max_row = ws.max_row
        
        # 遍历该列的所有单元格
        for row in range(min_row, max_row + 1):
            # 检查条件列是否有值
            if cond_col_letter:
                cond_cell = ws[f"{cond_col_letter}{row}"]
                # 使用新方法判断是否为有效数据行，而不是表头
                if not self._is_valid_data_cell(cond_cell):
                    continue
                    
            # 生成期望的公式
            expected_formula = tpl
            
            # 替换模板中的占位符
            pattern = r'\{([A-Za-z]+)\}'
            for match in re.finditer(pattern, tpl):
                col = match.group(1)
                expected_formula = expected_formula.replace(f"{{{col}}}", f"{col}{row}")
                
            # 验证单元格公式
            cell_ref = f"{col_letter}{row}"
            try:
                cell = ws[cell_ref]
                
                # 获取单元格的公式
                current_formula = cell.value if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') else None
                
                if current_formula is None and hasattr(cell, 'formula') and cell.formula:
                    current_formula = f"={cell.formula}" if not cell.formula.startswith('=') else cell.formula
                
                # 如果没有公式但应该有，或者公式不匹配
                if not current_formula or current_formula != expected_formula:
                    self.errors.append({
                        'file': file_path,
                        'sheet': rule.get('sheet', ''),
                        'cell': cell_ref,
                        'current_value': current_formula or "(无公式)",
                        'expected_value': expected_formula,
                        'rule_type': 'COL_FORMULA'
                    })
            except Exception as e:
                print(f"验证列公式 '{cell_ref}' 时出错: {str(e)}")

    def _validate_col_aggregate_check(self, wb, ws, rule, file_path):
        """验证列聚合检查规则 - 严格检查列中是否存在纯聚合函数（不包含额外运算）"""
        cell_ref = rule.get('cell_ref', '')
        agg_func = rule.get('FUNC', 'SUM').upper()  # 默认为SUM
        target_col = rule.get('TARGET_COL', '')
        cond_col = rule.get('COND_COL', '')
        
        # 如果没有指定目标列，则使用cell_ref中的列
        if not target_col:
            if re.match(r'^[A-Za-z]+列$', cell_ref) or re.match(r'^[A-Za-z]+$', cell_ref):
                target_col = re.match(r'^([A-Za-z]+)', cell_ref).group(1)
        
        # 如果仍然没有目标列，则返回
        if not target_col:
            return
        
        # 获取条件列
        cond_col_letter = cond_col if re.match(r'^[A-Za-z]+$', cond_col) else None
        
        # 检查条件列是否有有效数据（非表头）
        has_condition_data = False
        if cond_col_letter:
            for row in range(ws.min_row, ws.max_row + 1):
                cond_cell = ws[f"{cond_col_letter}{row}"]
                # 使用新方法判断是否为有效数据行
                if self._is_valid_data_cell(cond_cell):
                    has_condition_data = True
                    break
            
            if not has_condition_data:
                # 条件列没有有效数据，跳过此规则
                return
        
        # 检查目标列中是否存在纯聚合函数
        found_pure_aggregate = False
        invalid_formulas = []  # 记录不符合要求的公式

        for row in range(ws.min_row, ws.max_row + 1):
            cell = ws[f"{target_col}{row}"]

            # 获取单元格的公式
            current_formula = cell.value if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') else None

            if current_formula is None and hasattr(cell, 'formula') and cell.formula:
                current_formula = f"={cell.formula}" if not cell.formula.startswith('=') else cell.formula

            # 如果没有公式，跳过
            if not current_formula:
                continue

            # 使用严格验证检查是否为纯聚合函数
            if self._is_pure_aggregate_formula(current_formula, agg_func, target_col):
                found_pure_aggregate = True
                break
            else:
                # 检查是否包含聚合函数但不是纯函数（有额外运算）
                pattern = rf'{agg_func}\s*\(\s*{target_col}\d+:{target_col}\d+'
                if re.search(pattern, current_formula, re.IGNORECASE):
                    invalid_formulas.append({
                        'cell': f"{target_col}{row}",
                        'formula': current_formula
                    })
        
        # 如果没有找到符合条件的纯聚合函数，添加错误
        if not found_pure_aggregate:
            if invalid_formulas:
                # 如果有包含聚合函数但不纯的公式，报告具体错误
                for invalid in invalid_formulas:
                    self.errors.append({
                        'file': file_path,
                        'sheet': rule.get('sheet', ''),
                        'cell': invalid['cell'],
                        'current_value': invalid['formula'],
                        'expected_value': f"应为纯{agg_func}函数，如：={agg_func}({target_col}x:{target_col}y)，不允许额外运算",
                        'rule_type': 'COL_AGGREGATE_CHECK'
                    })
            else:
                # 如果完全没有找到聚合函数
                self.errors.append({
                    'file': file_path,
                    'sheet': rule.get('sheet', ''),
                    'cell': f"{target_col}列",
                    'current_value': "(未找到纯聚合函数)",
                    'expected_value': f"应包含纯{agg_func}({target_col}x:{target_col}y)格式的公式，不允许额外运算",
                    'rule_type': 'COL_AGGREGATE_CHECK'
                })


class RuleParser:
    """规则文件解析器"""
    
    def __init__(self, rule_file):
        self.rule_file = rule_file
        self.rules = []
        
    def parse_rules(self):
        """解析规则文件"""
        try:
            if self.rule_file.lower().endswith('.csv'):
                return self._parse_csv_rules()
            elif self.rule_file.lower().endswith('.xlsx'):
                return self._parse_excel_rules()
            elif self.rule_file.lower().endswith('.xls'):
                return self._parse_xls_rules()
            else:
                raise ValueError(f"不支持的规则文件格式: {self.rule_file}")
        except Exception as e:
            print(f"解析规则文件出错: {str(e)}")
            print(traceback.format_exc())
            return []
    
    def _extract_column_reference(self, cell_ref):
        """从复杂的列描述中提取核心列引用
        例如：从"P列（C列含"工日合计"的行）"中提取"P列"
        """
        # 如果描述包含括号，取括号前的部分
        if '（' in cell_ref:
            return cell_ref.split('（')[0].strip()
        if '(' in cell_ref:
            return cell_ref.split('(')[0].strip()
        # 否则返回原始引用
        return cell_ref.strip()
            
    def _parse_csv_rules(self):
        """解析CSV格式的规则文件"""
        import csv
        
        rules = []
        try:
            with open(self.rule_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                # 跳过标题行
                next(reader, None)
                
                for row in reader:
                    if len(row) < 3:  # 至少需要3列：子表名、单元格、期望值
                        continue
                        
                    sheet_name = row[0].strip()
                    raw_cell_ref = row[1].strip()
                    # 提取核心列引用
                    cell_ref = self._extract_column_reference(raw_cell_ref)
                    expected_value = row[2].strip() if len(row) > 2 else ""
                    dynamic_rule = row[3].strip() if len(row) > 3 else ""
                    
                    # 创建规则对象
                    rule = {
                        'sheet': sheet_name,
                        'cell_ref': cell_ref,
                        'raw_cell_ref': raw_cell_ref,  # 保存原始引用
                        'expected_value': expected_value,
                        'dynamic_rule': dynamic_rule
                    }
                    
                    # 解析动态规则
                    if dynamic_rule:
                        rule.update(self._parse_dynamic_rule(dynamic_rule))
                        
                    rules.append(rule)
        except Exception as e:
            print(f"解析CSV规则文件出错: {str(e)}")
            print(traceback.format_exc())
            
        return rules
    
    def _parse_xls_rules(self):
        """解析旧格式的.xls文件，使用xlrd库"""
        try:
            import xlrd
        except ModuleNotFoundError:
            print("错误: 未找到xlrd库，无法读取.xls格式的文件。")
            print("请通过运行以下命令安装xlrd库：")
            print("pip install xlrd==1.2.0")
            print("注意: 较新版本的xlrd可能不支持.xls格式，建议使用1.2.0版本。")
            return []
        
        rules = []
        try:
            # 打开工作簿
            wb = xlrd.open_workbook(self.rule_file)
            sheet = wb.sheet_by_index(0)  # 使用第一个工作表
            
            # 跳过标题行
            for row_idx in range(1, sheet.nrows):
                # 读取每一行的数据
                row_values = sheet.row_values(row_idx)
                
                if len(row_values) < 3:  # 至少需要3列
                    continue
                    
                sheet_name = str(row_values[0]).strip()
                raw_cell_ref = str(row_values[1]).strip()
                # 提取核心列引用
                cell_ref = self._extract_column_reference(raw_cell_ref)
                expected_value = str(row_values[2]).strip() if len(row_values) > 2 else ""
                dynamic_rule = str(row_values[3]).strip() if len(row_values) > 3 else ""
                
                if not sheet_name or not cell_ref:
                    continue
                    
                # 创建规则对象
                rule = {
                    'sheet': sheet_name,
                    'cell_ref': cell_ref,
                    'raw_cell_ref': raw_cell_ref,  # 保存原始引用
                    'expected_value': expected_value,
                    'dynamic_rule': dynamic_rule
                }
                
                # 解析动态规则
                if dynamic_rule:
                    rule.update(self._parse_dynamic_rule(dynamic_rule))
                    
                rules.append(rule)
                
        except Exception as e:
            print(f"解析XLS规则文件出错: {str(e)}")
            print(traceback.format_exc())
            
        return rules
        
    def _parse_excel_rules(self):
        """解析新格式的.xlsx文件，使用openpyxl库"""
        import openpyxl
        
        rules = []
        try:
            wb = openpyxl.load_workbook(self.rule_file, read_only=True, data_only=True)
            sheet = wb.active  # 使用第一个工作表
            
            # 跳过标题行
            for row in list(sheet.rows)[1:]:
                if len(row) < 3:  # 至少需要3列
                    continue
                    
                sheet_name = str(row[0].value).strip() if row[0].value else ""
                raw_cell_ref = str(row[1].value).strip() if row[1].value else ""
                # 提取核心列引用
                cell_ref = self._extract_column_reference(raw_cell_ref)
                expected_value = str(row[2].value).strip() if row[2].value else ""
                dynamic_rule = str(row[3].value).strip() if len(row) > 3 and row[3].value else ""
                
                if not sheet_name or not cell_ref:
                    continue
                    
                # 创建规则对象
                rule = {
                    'sheet': sheet_name,
                    'cell_ref': cell_ref,
                    'raw_cell_ref': raw_cell_ref,  # 保存原始引用
                    'expected_value': expected_value,
                    'dynamic_rule': dynamic_rule
                }
                
                # 解析动态规则
                if dynamic_rule:
                    rule.update(self._parse_dynamic_rule(dynamic_rule))
                    
                rules.append(rule)
                
        except Exception as e:
            print(f"解析Excel规则文件出错: {str(e)}")
            print(traceback.format_exc())
            
        return rules
        
    def _parse_dynamic_rule(self, dynamic_rule):
        """解析D列的动态规则"""
        rule_data = {}
        
        # 如果没有动态规则，返回空字典
        if not dynamic_rule:
            return rule_data
            
        # 将规则拆分为键值对
        pairs = dynamic_rule.split(';')
        for pair in pairs:
            pair = pair.strip()
            if not pair:
                continue
                
            # 解析键值对
            if ':' in pair:
                key, value = pair.split(':', 1)
                key = key.strip()
                value = value.strip()
                
                # 处理引号包裹的值
                if value.startswith('"') and value.endswith('"'):
                    value = value[1:-1]
                    
                rule_data[key] = value
        
        # 设置规则类型
        if 'TYPE' in rule_data:
            rule_data['rule_type'] = rule_data['TYPE']
        else:
            rule_data['rule_type'] = 'FIXED_VALUE'  # 默认为固定值规则
            
        return rule_data


if __name__ == "__main__":
    app = AuditTool()
    app.mainloop()


